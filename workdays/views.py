import calendar as _cal
import io
import logging
import os
from datetime import timedelta, datetime as _datetime, time as _time, date as _date
from django.conf import settings
from django.db import models
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from core.permissions import IsInternalRequest
from employees.models import Employee
from authentication.serializers import AGENT_ACTIVE_THRESHOLD_MINUTES
from .models import Workday, DailyReport, CaptureConfig, InactivityPeriod, ExecutiveMessage, CalendarNote, EmployeeLeave, ActivityItem

logger = logging.getLogger(__name__)


class WorkdayStartView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({'error': 'Perfil de empleado no encontrado'}, status=404)

        if Workday.objects.filter(employee=employee, status=Workday.STATUS_IN_PROGRESS).exists():
            return Response(
                {'error': 'Ya tienes una jornada activa'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lat = request.data.get('latitude')
        lng = request.data.get('longitude')

        if employee.mobile_type == Employee.MOBILE_TYPE_GPS and (not lat or not lng):
            return Response(
                {'error': 'La ubicación es requerida para iniciar la jornada', 'location_required': True},
                status=status.HTTP_400_BAD_REQUEST,
            )

        workday = Workday.objects.create(
            employee=employee,
            start_time=timezone.now(),
            status=Workday.STATUS_IN_PROGRESS,
            start_latitude=lat or None,
            start_longitude=lng or None,
        )
        return Response({
            'workday_id': workday.id,
            'start_time': workday.start_time,
            'status': workday.status,
        }, status=status.HTTP_201_CREATED)


class WorkdayEndView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({'error': 'Perfil de empleado no encontrado'}, status=404)

        workday_id = request.data.get('workday_id')
        activities_done = request.data.get('activities_done', '').strip()
        activities_planned = request.data.get('activities_planned', '').strip()
        lat = request.data.get('latitude')
        lng = request.data.get('longitude')

        if not workday_id:
            return Response({'error': 'workday_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        mobile_type = employee.mobile_type
        if mobile_type == Employee.MOBILE_TYPE_GPS and (not lat or not lng):
            return Response(
                {'error': 'La ubicación es requerida para finalizar la jornada', 'location_required': True},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if mobile_type == Employee.MOBILE_TYPE_REPORT and (not activities_done or not activities_planned):
            return Response(
                {'error': 'Debes completar el reporte diario antes de finalizar', 'report_required': True},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workday = Workday.objects.get(id=workday_id, employee=employee, status=Workday.STATUS_IN_PROGRESS)
        except Workday.DoesNotExist:
            return Response(
                {'error': 'Jornada activa no encontrada'},
                status=status.HTTP_404_NOT_FOUND,
            )

        end_time = timezone.now()
        duration = int((end_time - workday.start_time).total_seconds() // 60)
        workday.end_time = end_time
        workday.duration_minutes = duration
        workday.status = Workday.STATUS_COMPLETED
        workday.end_latitude  = lat or None
        workday.end_longitude = lng or None
        workday.save()

        # Captura de pantalla al finalizar jornada
        try:
            from asgiref.sync import async_to_sync
            from channels.layers import get_channel_layer
            async_to_sync(get_channel_layer().group_send)(
                f'agent_{employee.id}',
                {'type': 'capture_command', 'command': 'capture'},
            )
        except Exception:
            pass

        report = DailyReport.objects.create(
            workday=workday,
            activities_done=activities_done,
            activities_planned=activities_planned,
        )

        # Clasificar actividades en categorías (nunca bloquear el fin de jornada).
        # Se omite el almacén (rol "otro"): las estadísticas son solo Supervisores vs PMs.
        if employee.role != Employee.ROLE_OTRO:
            try:
                from .classifier import sync_report_items
                sync_report_items(report)
            except Exception:
                logger.exception('Fallo clasificando reporte %s', report.id)

        return Response({
            'workday_id': workday.id,
            'duration_minutes': duration,
            'status': workday.status,
        })


class ActiveWorkdayView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({'active': False})

        now = timezone.now()

        # Intervalo efectivo: override por empleado o global
        effective_interval = (
            employee.capture_interval_minutes
            if employee.capture_interval_minutes is not None
            else CaptureConfig.get().capture_interval_minutes
        )

        # Detectar inactividad: si el gap desde el último heartbeat supera 1.5× el intervalo
        # y hay una jornada activa, registrar el período de inactividad
        prev_seen = employee.agent_last_seen
        gap_threshold = timedelta(minutes=effective_interval * 1.5)
        if prev_seen and (now - prev_seen) > gap_threshold:
            try:
                active_workday = Workday.objects.filter(employee=employee, status=Workday.STATUS_IN_PROGRESS).order_by('-start_time').first()
                if not active_workday: raise Workday.DoesNotExist
                gap_minutes = int((now - prev_seen).total_seconds() // 60)
                InactivityPeriod.objects.create(
                    workday=active_workday,
                    started_at=prev_seen,
                    ended_at=now,
                    duration_minutes=gap_minutes,
                )
            except Workday.DoesNotExist:
                pass

        # Stamp agent heartbeat y versión en cada poll
        employee.agent_last_seen = now
        version = request.headers.get('X-Agent-Version', '')
        update_fields = ['agent_last_seen']
        if version and employee.agent_version != version:
            employee.agent_version = version
            update_fields.append('agent_version')
        employee.save(update_fields=update_fields)

        screenshots_enabled = employee.screenshots_enabled

        try:
            workday = Workday.objects.filter(employee=employee, status=Workday.STATUS_IN_PROGRESS).order_by('-start_time').first()
            if not workday: raise Workday.DoesNotExist
            inactive_minutes = InactivityPeriod.objects.filter(workday=workday).aggregate(
                total=models.Sum('duration_minutes')
            )['total'] or 0
            return Response({
                'active': True,
                'workday_id': workday.id,
                'start_time': workday.start_time,
                'capture_interval_minutes': effective_interval,
                'screenshots_enabled': screenshots_enabled,
                'inactive_minutes': inactive_minutes,
            })
        except Workday.DoesNotExist:
            return Response({
                'active': False,
                'capture_interval_minutes': effective_interval,
                'screenshots_enabled': screenshots_enabled,
            })


def _local_now():
    return timezone.localtime(timezone.now())


def _month_range(year, month):
    """Devuelve (first_aware, last_aware) en zona local para filtrar por mes."""
    first = timezone.make_aware(_datetime(_date(year, month, 1).year, month, 1, 0, 0, 0))
    last_day = _cal.monthrange(year, month)[1]
    last = timezone.make_aware(_datetime(year, month, last_day, 23, 59, 59))
    return first, last


class CloseStaleWorkdaysView(APIView):
    """Endpoint interno para n8n: cierra jornadas abiertas del día (corre a las 23:59)."""
    permission_classes = [IsInternalRequest]

    def post(self, request):
        closed = _close_stale_workdays()
        return Response({'closed': closed})


class CloseDayMessagesView(APIView):
    """Endpoint interno para n8n: marca mensajes no leídos del día ANTERIOR con day_closed
    (corre a las 00:01 del día siguiente)."""
    permission_classes = [IsInternalRequest]

    def post(self, request):
        target = _local_now().date() - timedelta(days=1)
        day_start = timezone.make_aware(_datetime.combine(target, _time(0, 0)))
        day_end   = timezone.make_aware(_datetime.combine(target, _time(23, 59, 59)))
        flagged = ExecutiveMessage.objects.filter(
            acknowledged_at__isnull=True,
            day_closed=False,
            sent_at__range=(day_start, day_end),
        ).update(day_closed=True)
        return Response({'flagged': flagged})


class DoneActivitiesView(APIView):
    """Endpoint interno para n8n (cron 00:05): devuelve las actividades 'done' de una
    fecha (default: día ANTERIOR), agrupadas por empleado, junto con la lista DINÁMICA
    de categorías activas para que n8n arme el prompt del LLM.

    `?date=YYYY-MM-DD` permite elegir la fecha (para pruebas). Solo ítems sin
    `manual_override` (los manuales no se tocan)."""
    permission_classes = [IsInternalRequest]

    def get(self, request):
        from .models import ActivityCategory, ActivityTag

        date_param = request.query_params.get('date', '')
        target = _local_now().date() - timedelta(days=1)
        if date_param:
            try:
                target = _date.fromisoformat(date_param)
            except ValueError:
                pass
        day_start = timezone.make_aware(_datetime.combine(target, _time(0, 0)))
        day_end   = timezone.make_aware(_datetime.combine(target, _time(23, 59, 59)))

        # Solo Supervisores y Project Managers: el LLM nunca clasifica a otros roles
        # (almacén/administrativos), que además quedan fuera de las estadísticas.
        # `role` se deriva de `cargo` en Python, así que se resuelven los IDs elegibles.
        eligible_ids = [
            e.id for e in Employee.objects.filter(is_active=True, is_executive=False)
            if e.role in (Employee.ROLE_SUPERVISOR, Employee.ROLE_PROJECT_MANAGER)]
        items = (ActivityItem.objects
                 .filter(kind=ActivityItem.KIND_DONE, manual_override=False,
                         report__workday__employee_id__in=eligible_ids,
                         report__workday__start_time__range=(day_start, day_end))
                 .select_related('report__workday__employee')
                 .order_by('report__workday__employee_id', 'position'))

        by_emp = {}
        for it in items:
            emp = it.report.workday.employee
            bucket = by_emp.setdefault(emp.id, {
                'employee_id': emp.id, 'employee_name': emp.full_name, 'items': []})
            bucket['items'].append({
                'id': it.id, 'text': it.text, 'fallback_category': it.category})

        categories = [
            {'code': c.code, 'label': c.label}
            for c in ActivityCategory.objects.filter(is_active=True).order_by('order', 'label')
        ]
        # Tags canónicos existentes (proyecto/ubicación) para que el LLM reutilice
        # nombres. Proyectos: oficiales (con código) primero. Cap acotado para no
        # inflar el prompt (se repite por cada ítem → ejecuciones grandes en n8n).
        known_tags = {
            ActivityTag.KIND_PROJECT: list(
                ActivityTag.objects.filter(kind=ActivityTag.KIND_PROJECT,
                                           canonical__isnull=True, ignored=False)
                .order_by('-code', 'name').values_list('name', flat=True)[:80]),
            ActivityTag.KIND_LOCATION: list(
                ActivityTag.objects.filter(kind=ActivityTag.KIND_LOCATION,
                                           canonical__isnull=True, ignored=False)
                .order_by('name').values_list('name', flat=True)[:40]),
        }
        return Response({
            'date': target.isoformat(),
            'categories': categories,
            'known_tags': known_tags,
            'employees': list(by_emp.values()),
        })


class ClassifyDoneView(APIView):
    """Endpoint interno para n8n: aplica la clasificación LLM a ítems 'done'.

    Body: {"results": [{"id": <activity_item_id>, "category": "<code>"}, ...]}.
    El LLM manda, pero el determinístico (ya corrió al cerrar la jornada) queda como
    fallback: categoría inválida/inactiva o ítem inexistente/override se ignoran y se
    conserva la categoría previa."""
    permission_classes = [IsInternalRequest]

    def post(self, request):
        from .models import ActivityCategory, ActivityTag, SEDE_NAMES

        results = request.data.get('results') or []
        valid = set(ActivityCategory.objects.filter(is_active=True).values_list('code', flat=True))
        # La UBICACIÓN no sale del texto (el LLM la extrae pero la ignoramos): es la
        # SEDE del empleado que reportó (employee.ciudad). Proyecto y entregable sí del LLM.
        kind_fields = [
            ('projects',     ActivityTag.KIND_PROJECT),
            ('deliverables', ActivityTag.KIND_DELIVERABLE),
        ]
        updated = skipped = 0
        for r in results:
            if not isinstance(r, dict):
                skipped += 1
                continue
            code = r.get('category')
            try:
                item_id = int(r.get('id'))
            except (TypeError, ValueError):
                skipped += 1
                continue
            if code not in valid:
                skipped += 1
                continue
            try:
                item = (ActivityItem.objects
                        .select_related('report__workday__employee')
                        .get(id=item_id, kind=ActivityItem.KIND_DONE, manual_override=False))
            except ActivityItem.DoesNotExist:
                skipped += 1
                continue
            item.category = code
            item.source = ActivityItem.SOURCE_LLM
            item.matched_keyword = ''
            item.save(update_fields=['category', 'source', 'matched_keyword', 'classified_at'])
            # Tags: proyecto/entregable del LLM + ubicación = sede del empleado.
            tag_objs = []
            sede = (item.report.workday.employee.ciudad or '').strip().upper()
            if sede and sede != 'NONE':
                loc = ActivityTag.resolve(ActivityTag.KIND_LOCATION, SEDE_NAMES.get(sede, sede))
                if loc:
                    tag_objs.append(loc)
            for field, kind in kind_fields:
                for name in (r.get(field) or []):
                    tag = ActivityTag.resolve(kind, name if isinstance(name, str) else str(name))
                    if tag:
                        tag_objs.append(tag)
            item.tags.set(tag_objs)
            updated += 1
        return Response({'updated': updated, 'skipped': skipped})


class TagsListView(APIView):
    """Lista de tags por `kind` con conteo de ítems (incluye los de sus alias).
    Solo ejecutivos. Para el panel de fusión del dashboard."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from collections import defaultdict
        from django.db.models import Count
        from .models import ActivityTag
        try:
            emp = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (emp.is_executive or emp.can_edit_tags):
            return Response({'error': 'Acceso no autorizado'}, status=403)

        kind = request.query_params.get('kind', ActivityTag.KIND_PROJECT)
        tags = list(ActivityTag.objects.filter(kind=kind, ignored=False))
        per_tag = dict(ActivityTag.objects.filter(kind=kind, ignored=False)
                       .annotate(c=Count('items')).values_list('id', 'c'))
        agg = {t.id: {'id': t.id, 'name': t.name, 'code': t.code,
                      'count': 0, 'aliases': 0, 'alias_list': [], 'employees': []}
               for t in tags if t.canonical_id is None}
        for t in tags:
            root = t.canonical_id or t.id
            if root in agg:
                agg[root]['count'] += per_tag.get(t.id, 0)
                if t.canonical_id is not None:
                    agg[root]['aliases'] += 1
                    agg[root]['alias_list'].append({'id': t.id, 'name': t.name})

        # Empleados por tag (agregando alias en su canónico), en una sola query.
        emp_by_canon = defaultdict(lambda: defaultdict(int))
        people = {}                        # emp_id -> name (para el filtro)
        per_emp_canon = defaultdict(set)   # emp_id -> {canonical tag ids}
        emp_rows = (ActivityItem.objects
                    .filter(tags__kind=kind, tags__ignored=False)
                    .values('tags__id', 'tags__canonical_id',
                            'report__workday__employee_id',
                            'report__workday__employee__full_name')
                    .annotate(n=Count('id', distinct=True)))
        for r in emp_rows:
            cid = r['tags__canonical_id'] or r['tags__id']
            eid = r['report__workday__employee_id']
            name = r['report__workday__employee__full_name']
            if cid in agg and name:
                emp_by_canon[cid][name] += r['n']
                people[eid] = name
                per_emp_canon[eid].add(cid)
        for cid, v in agg.items():
            emps = sorted(emp_by_canon.get(cid, {}).items(), key=lambda x: (-x[1], x[0].lower()))
            v['employees'] = [{'name': n, 'count': c} for n, c in emps]

        # Mostrar tags con ítems, o los oficiales (con `code`) aunque tengan 0 — son
        # el destino para asociar variantes. Se ocultan los huérfanos sin código ni ítems.
        out = sorted((v for v in agg.values() if v['count'] > 0 or v['code']),
                     key=lambda x: (-x['count'], x['name'].lower()))
        # Filtro opcional por empleado: solo los tags donde ese empleado trabajó.
        emp_filter = request.query_params.get('employee') or ''
        if emp_filter:
            try:
                allowed = per_emp_canon.get(int(emp_filter), set())
            except (TypeError, ValueError):
                allowed = set()
            out = [v for v in out if v['id'] in allowed]
        all_employees = [{'id': i, 'name': n}
                         for i, n in sorted(people.items(), key=lambda x: x[1].lower())]
        return Response({'kind': kind, 'tags': out, 'all_employees': all_employees})


class TagMergeView(APIView):
    """Fusiona variantes en un tag canónico (no destructivo): `from` (y sus alias)
    pasan a apuntar a `into`. Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from .models import ActivityTag
        try:
            emp = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (emp.is_executive or emp.can_edit_tags):
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            into = ActivityTag.objects.get(id=request.data.get('into'))
        except (ActivityTag.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'into inválido'}, status=400)
        into = into.root()  # nunca fusionar hacia un alias
        merged = 0
        for fid in (request.data.get('from') or []):
            try:
                t = ActivityTag.objects.get(id=fid, kind=into.kind)
            except (ActivityTag.DoesNotExist, ValueError, TypeError):
                continue
            if t.id == into.id:
                continue
            t.aliases.update(canonical=into)   # re-enraizar alias previos de t
            t.canonical = into
            t.save(update_fields=['canonical'])
            merged += 1
        return Response({'into': into.id, 'name': into.name, 'merged': merged})


class TagUnmergeView(APIView):
    """Desagrupa: los tags indicados (que son alias) vuelven a ser canónicos
    independientes (`canonical=None`). Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from .models import ActivityTag
        try:
            emp = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (emp.is_executive or emp.can_edit_tags):
            return Response({'error': 'Acceso no autorizado'}, status=403)
        ids = [i for i in (request.data.get('ids') or [])
               if isinstance(i, int) or (isinstance(i, str) and i.isdigit())]
        n = (ActivityTag.objects.filter(id__in=ids, canonical__isnull=False)
             .update(canonical=None))
        return Response({'unmerged': n})


class TagIgnoreView(APIView):
    """Descarta tags mal extraídos (ej. nombres de persona tomados como proyecto).
    Los marca `ignored`, desvincula sus ítems y evita que se vuelvan a asociar. Solo
    ejecutivos."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from .models import ActivityTag
        try:
            emp = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (emp.is_executive or emp.can_edit_tags):
            return Response({'error': 'Acceso no autorizado'}, status=403)

        ids = [i for i in (request.data.get('ids') or [])
               if isinstance(i, int) or (isinstance(i, str) and i.isdigit())]
        n = 0
        for t in ActivityTag.objects.filter(id__in=ids):
            family = [t] + list(t.aliases.all())   # el tag y todos sus alias
            for ft in family:
                ft.items.clear()                   # desaparece de reportes
                if not ft.ignored:
                    ft.ignored = True
                    ft.save(update_fields=['ignored'])
            n += 1
        return Response({'ignored': n})


class TagRenameView(APIView):
    """Edita el nombre visible de un tag (no cambia su `key` de deduplicación, así
    futuras extracciones del nombre original siguen ruteando aquí). Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def post(self, request, tag_id):
        from .models import ActivityTag
        try:
            emp = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (emp.is_executive or emp.can_edit_tags):
            return Response({'error': 'Acceso no autorizado'}, status=403)
        try:
            tag = ActivityTag.objects.get(id=tag_id)
        except (ActivityTag.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'no encontrado'}, status=404)
        name = (request.data.get('name') or '').strip()[:120]
        if not name:
            return Response({'error': 'nombre vacío'}, status=400)
        tag.name = name
        tag.save(update_fields=['name'])
        return Response({'id': tag.id, 'name': tag.name})


class TagEmployeesView(APIView):
    """Empleados que trabajaron en un tag (proyecto/ubicación/entregable), incluyendo
    los ítems de sus alias. Para el cruce tag↔empleado en el dashboard. Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def get(self, request, tag_id):
        from django.db.models import Count
        from .models import ActivityTag
        try:
            emp = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (emp.is_executive or emp.can_edit_tags):
            return Response({'error': 'Acceso no autorizado'}, status=403)
        try:
            tag = ActivityTag.objects.get(id=tag_id)
        except (ActivityTag.DoesNotExist, ValueError, TypeError):
            return Response({'error': 'no encontrado'}, status=404)
        tag = tag.root()
        ids = [tag.id] + list(tag.aliases.values_list('id', flat=True))
        rows = (ActivityItem.objects.filter(tags__in=ids)
                .values('report__workday__employee_id',
                        'report__workday__employee__full_name')
                .annotate(n=Count('id', distinct=True)).order_by('-n'))
        employees = [{'id': r['report__workday__employee_id'],
                      'name': r['report__workday__employee__full_name'], 'count': r['n']}
                     for r in rows]
        return Response({'tag': tag.name, 'employees': employees})


def _close_stale_workdays():
    """Cierra jornadas in_progress que el empleado no cerró del día ANTERIOR (corre a las 00:00
    del día siguiente). Las marca como INCOMPLETE + auto_closed y avisa al empleado vía skyBot."""
    target = _local_now().date() - timedelta(days=1)
    day_start = timezone.make_aware(_datetime.combine(target, _time(0, 0)))
    day_end   = timezone.make_aware(_datetime.combine(target, _time(23, 59, 59)))
    stale = Workday.objects.filter(
        status=Workday.STATUS_IN_PROGRESS,
        start_time__range=(day_start, day_end),
    ).select_related('employee')
    count = 0
    for w in stale:
        local_start = timezone.localtime(w.start_time)
        close_at = timezone.make_aware(
            _datetime.combine(local_start.date(), _time(17, 0))
        )
        if w.start_time >= close_at:
            close_at = w.start_time + timedelta(minutes=1)
        duration = max(1, int((close_at - w.start_time).total_seconds() // 60))
        w.end_time = close_at
        w.duration_minutes = duration
        w.status = Workday.STATUS_INCOMPLETE
        w.auto_closed = True
        w.save(update_fields=['end_time', 'duration_minutes', 'status', 'auto_closed'])

        # Aviso automático al empleado (remitente: bot del sistema skyBot → sender=None)
        fecha = local_start.strftime('%d/%m/%Y')
        ExecutiveMessage.objects.create(
            sender=None,
            recipient=w.employee,
            body=f'Tu jornada del {fecha} quedó abierta y fue cerrada automáticamente a las 17:00.',
        )
        count += 1
    return count


class EmployeeOverviewView(APIView):
    """Vista exclusiva para ejecutivos: estado de todos los empleados."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not employee.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        # Fecha solicitada (por defecto hoy)
        date_param = request.query_params.get('date', '')
        today = _local_now().date()
        try:
            target_date = _date.fromisoformat(date_param) if date_param else today
        except ValueError:
            target_date = today
        is_today = (target_date == today)


        now = timezone.now()
        employees = Employee.objects.filter(is_active=True, is_executive=False).order_by('full_name')
        global_interval = CaptureConfig.get().capture_interval_minutes

        target_start = timezone.make_aware(_datetime.combine(target_date, _time(0, 0)))
        target_end   = timezone.make_aware(_datetime.combine(target_date, _time(23, 59, 59)))

        # Mensajes del día visualizado (confirmados o no) agrupados por destinatario
        messages_by_employee = {}
        for m in ExecutiveMessage.objects.filter(
            recipient__in=employees,
            sent_at__range=(target_start, target_end),
        ).order_by('sent_at').select_related('sender'):
            messages_by_employee.setdefault(m.recipient_id, []).append(m)

        # Jornadas activas (in_progress) iniciadas en el día visualizado.
        # Una jornada que el empleado dejó abierta en un día pasado se sigue
        # mostrando como activa en ese día (hasta que el cierre de las 00:00 la tome).
        active_workdays = {
            w.employee_id: w
            for w in Workday.objects.filter(
                employee__in=employees,
                status=Workday.STATUS_IN_PROGRESS,
                start_time__range=(target_start, target_end),
            ).select_related('daily_report')
        }
        # La inactividad en vivo solo aplica al día de hoy.
        inactive_by_workday = {}
        if is_today:
            inactive_by_workday = {
                row['workday_id']: row['total']
                for row in InactivityPeriod.objects.filter(
                    workday_id__in=active_workdays.values()
                ).values('workday_id').annotate(total=models.Sum('duration_minutes'))
            }

        # Jornadas cerradas que INICIARON el día solicitado: completadas por el
        # empleado o auto-cerradas sin finalizar (INCOMPLETE). Se filtra por
        # start_time (la jornada pertenece a su día de inicio), evitando que una
        # jornada que se extendió de un día a otro aparezca en el día equivocado.
        completed_workdays = {
            w.employee_id: w
            for w in Workday.objects.filter(
                employee__in=employees,
                status__in=[Workday.STATUS_COMPLETED, Workday.STATUS_INCOMPLETE],
                start_time__range=(target_start, target_end),
            ).select_related('daily_report')
        }

        result = []
        for emp in employees:
            workday  = active_workdays.get(emp.id)
            completed = completed_workdays.get(emp.id)

            effective_interval = (
                emp.capture_interval_minutes
                if emp.capture_interval_minutes is not None
                else global_interval
            )

            agent_active = emp.agent_online if is_today else False
            emp_msgs = messages_by_employee.get(emp.id, [])
            result.append({
                'id': emp.id,
                'full_name': emp.full_name,
                'messages': [
                    {'body': m.body, 'sender': m.sender.full_name if m.sender else 'skyBot', 'acknowledged': bool(m.acknowledged_at)}
                    for m in emp_msgs
                ],
                'solo_movil': emp.solo_movil,
                'mobile_device_bound': bool(emp.mobile_device_id),
                'agent_is_active': agent_active,
                'agent_version': emp.agent_version,
                'agent_last_seen': emp.agent_last_seen,
                'capture_interval_minutes': effective_interval,
                'screenshots_enabled': emp.screenshots_enabled,
                'skylog_access': emp.skylog_access,
                'workday': {
                    'active': True,
                    'workday_id': workday.id,
                    'start_time': workday.start_time,
                    'duration_minutes': int((now - workday.start_time).total_seconds() // 60),
                    'inactive_minutes': inactive_by_workday.get(workday.id, 0),
                    'auto_closed': workday.auto_closed,
                    'start_latitude':  float(workday.start_latitude)  if workday.start_latitude  else None,
                    'start_longitude': float(workday.start_longitude) if workday.start_longitude else None,
                    'activities_done': getattr(workday, 'daily_report', None) and workday.daily_report.activities_done or '',
                    'activities_planned': getattr(workday, 'daily_report', None) and workday.daily_report.activities_planned or '',
                } if workday else ({
                    'active': False,
                    'completed': True,
                    'start_time': completed.start_time,
                    'end_time': completed.end_time,
                    'duration_minutes': completed.duration_minutes,
                    'auto_closed': completed.auto_closed,
                    'activities_done': getattr(completed, 'daily_report', None) and completed.daily_report.activities_done or '',
                    'activities_planned': getattr(completed, 'daily_report', None) and completed.daily_report.activities_planned or '',
                } if completed else None),
            })

        agents_online = sum(1 for e in result if e['agent_is_active'])
        activities_count = ActivityItem.objects.filter(
            kind=ActivityItem.KIND_DONE,
            report__workday__start_time__range=(target_start, target_end),
            report__workday__employee__is_active=True,
            report__workday__employee__is_executive=False,
        ).count()
        return Response({
            'is_today': is_today,
            'summary': {
                'active_now': len(active_workdays),
                'completed_today': len(completed_workdays),
                'agents_online': agents_online,
                'total_employees': len(result),
                'activities': activities_count,
            },
            'employees': result,
        })


class CaptureNowView(APIView):
    """Ejecutivo solicita captura inmediata del agente de un empleado via WebSocket."""
    permission_classes = [IsAuthenticated]

    def post(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'agent_{employee_id}',
                {'type': 'capture_command', 'command': 'capture'},
            )
        except Exception as e:
            return Response({'error': f'Error al enviar comando al agente: {e}'}, status=500)
        return Response({'status': 'ok'})


class EmployeeSkylogToggleView(APIView):
    """Ejecutivo habilita/deshabilita el acceso a Skylog de un empleado."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        enabled = request.data.get('skylog_access')
        if enabled is None:
            return Response({'error': 'skylog_access es requerido'}, status=400)

        try:
            emp = Employee.objects.get(id=employee_id, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        emp.skylog_access = bool(enabled)
        emp.save(update_fields=['skylog_access'])
        return Response({'skylog_access': emp.skylog_access})


class EmployeeScreenshotsToggleView(APIView):
    """Ejecutivo habilita/deshabilita capturas de pantalla de un empleado."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        enabled = request.data.get('screenshots_enabled')
        if enabled is None:
            return Response({'error': 'screenshots_enabled es requerido'}, status=400)

        try:
            emp = Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        emp.screenshots_enabled = bool(enabled)
        emp.save(update_fields=['screenshots_enabled'])
        return Response({'screenshots_enabled': emp.screenshots_enabled})


class LastReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({})

        last = (
            Workday.objects
            .filter(employee=employee, status=Workday.STATUS_COMPLETED)
            .select_related('daily_report')
            .order_by('-end_time')
            .first()
        )

        if not last or not hasattr(last, 'daily_report'):
            return Response({'has_report': False})

        return Response({
            'has_report': True,
            'activities_done': last.daily_report.activities_done,
            'activities_planned': last.daily_report.activities_planned,
            'date': last.end_time.date(),
        })


class SendMessageView(APIView):
    """Ejecutivo envía un mensaje a un empleado específico."""
    permission_classes = [IsAuthenticated]

    def post(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        body = request.data.get('body', '').strip()
        if not body:
            return Response({'error': 'El mensaje no puede estar vacío'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            recipient = Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        message = ExecutiveMessage.objects.create(
            sender=executive,
            recipient=recipient,
            body=body,
        )

        # Notificar en tiempo real al dashboard del empleado (si está conectado)
        try:
            from asgiref.sync import async_to_sync
            from channels.layers import get_channel_layer
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'dashboard_{recipient.id}',
                {'type': 'new_message'},
            )
        except Exception:
            pass

        return Response({'id': message.id, 'sent_at': message.sent_at}, status=status.HTTP_201_CREATED)


class MessageLeadsView(APIView):
    """Usuario habilitado (can_message_leads) envía un mensaje a Supervisores y/o
    Project Managers no ejecutivos: por grupo (rol) o a una lista de personas.
    GET devuelve los destinatarios elegibles para el modo 'personas'."""
    permission_classes = [IsAuthenticated]

    def _check(self, request):
        try:
            sender = request.user.employee
        except Exception:
            return None, Response({'error': 'Perfil no encontrado'}, status=404)
        if not (sender.is_executive or sender.can_message_leads or request.user.is_superuser):
            return None, Response({'error': 'Acceso no autorizado'}, status=403)
        return sender, None

    def get(self, request):
        sender, err = self._check(request)
        if err:
            return err
        valid = {Employee.ROLE_SUPERVISOR, Employee.ROLE_PROJECT_MANAGER}
        people = [
            {'id': e.id, 'name': e.full_name, 'role': e.role, 'role_label': e.role_label}
            for e in Employee.objects.filter(is_active=True, is_executive=False).order_by('full_name')
            if e.role in valid and e.id != sender.id
        ]
        return Response(people)

    def post(self, request):
        sender, err = self._check(request)
        if err:
            return err

        body = request.data.get('body', '').strip()
        if not body:
            return Response({'error': 'El mensaje no puede estar vacío'}, status=status.HTTP_400_BAD_REQUEST)

        valid = {Employee.ROLE_SUPERVISOR, Employee.ROLE_PROJECT_MANAGER}
        recipient_ids = request.data.get('recipient_ids')

        # `role` es una property derivada de `cargo`, no una columna: filtrar en Python.
        if recipient_ids:
            # Modo 'personas': destinatarios específicos (validados como Sup/PM no ejecutivos).
            try:
                ids = {int(x) for x in recipient_ids}
            except (TypeError, ValueError):
                return Response({'error': 'recipient_ids inválido'}, status=status.HTTP_400_BAD_REQUEST)
            recipients = [
                emp for emp in Employee.objects.filter(id__in=ids, is_active=True, is_executive=False)
                if emp.role in valid and emp.id != sender.id
            ]
        else:
            # Modo 'grupo': todos los de los roles elegidos.
            targets = request.data.get('targets') or list(valid)
            targets = [t for t in targets if t in valid]
            if not targets:
                return Response({'error': 'Selecciona al menos un destino (Supervisores o PMs)'}, status=status.HTTP_400_BAD_REQUEST)
            recipients = [
                emp for emp in Employee.objects.filter(is_active=True, is_executive=False)
                if emp.role in targets and emp.id != sender.id
            ]

        if not recipients:
            return Response({'error': 'No hay destinatarios para el envío'}, status=status.HTTP_400_BAD_REQUEST)

        # Firma del remitente al pie del mensaje.
        signed_body = f"{body}\n\n— {sender.full_name}"

        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer
        channel_layer = get_channel_layer()

        def _notify(emp_id):
            try:
                async_to_sync(channel_layer.group_send)(
                    f'dashboard_{emp_id}', {'type': 'new_message'},
                )
            except Exception:
                pass

        for emp in recipients:
            ExecutiveMessage.objects.create(sender=sender, recipient=emp, body=signed_body)
            _notify(emp.id)

        # El remitente también recibe una copia de su propio aviso (se muestra en verde).
        ExecutiveMessage.objects.create(sender=sender, recipient=sender, body=signed_body)
        _notify(sender.id)

        return Response({'sent': len(recipients)}, status=status.HTTP_201_CREATED)


class AdminPermissionsListView(APIView):
    """Superuser: lista de empleados activos con sus permisos granulares."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        employees = Employee.objects.filter(is_active=True).order_by('full_name')
        data = [{
            'id': e.id,
            'full_name': e.full_name,
            'cargo': e.cargo,
            'role': e.role,
            'role_label': e.role_label,
            'is_executive': e.is_executive,
            'can_message_leads': e.can_message_leads,
            'can_view_stats': e.can_view_stats,
            'can_edit_tags': e.can_edit_tags,
            'can_view_report': e.can_view_report,
        } for e in employees]
        return Response(data)


class AdminPermissionsUpdateView(APIView):
    """Superuser: habilita/deshabilita un permiso granular de un empleado."""
    permission_classes = [IsAuthenticated]
    ALLOWED_FIELDS = {'can_message_leads', 'can_view_stats', 'can_edit_tags', 'can_view_report'}

    def post(self, request, employee_id):
        if not request.user.is_superuser:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        field = request.data.get('field')
        if field not in self.ALLOWED_FIELDS:
            return Response({'error': 'Campo inválido'}, status=status.HTTP_400_BAD_REQUEST)
        value = bool(request.data.get('value'))

        try:
            emp = Employee.objects.get(id=employee_id, is_active=True)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        setattr(emp, field, value)
        emp.save(update_fields=[field])
        return Response({'id': emp.id, field: value})


class EmployeePendingMessagesView(APIView):
    """Ejecutivo obtiene los mensajes pendientes de un empleado específico."""
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            recipient = Employee.objects.get(id=employee_id, is_active=True)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        cutoff = timezone.now() - timedelta(hours=24)
        messages = (
            ExecutiveMessage.objects
            .filter(recipient=recipient, acknowledged_at__isnull=True, sent_at__gte=cutoff)
            .select_related('sender')
        )
        data = [
            {
                'id': m.id,
                'body': m.body,
                'sent_at': m.sent_at,
                'sender_name': m.sender.full_name if m.sender else 'skyBot',
            }
            for m in messages
        ]
        return Response(data)


class EmployeeViewDataView(APIView):
    """Ejecutivo obtiene todos los datos del dashboard de un empleado específico."""
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            emp = Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        now = timezone.now()

        profile = {
            'id': emp.id,
            'full_name': emp.full_name,
            'email': emp.user.email if emp.user else '',
            'agent_version': emp.agent_version,
            'agent_last_seen': emp.agent_last_seen,
            'agent_is_active': emp.agent_online,
            'solo_movil': emp.solo_movil,
            'skylog_access': emp.skylog_access,
            'is_executive': False,
        }

        workday_data = None
        wd = (
            Workday.objects.select_related('daily_report')
            .filter(employee=emp, status=Workday.STATUS_IN_PROGRESS)
            .order_by('-start_time')
            .first()
        )
        if wd:
            workday_data = {
                'active': True,
                'workday_id': wd.id,
                'start_time': wd.start_time,
                'duration_minutes': int((now - wd.start_time).total_seconds() // 60),
                'activities_done': getattr(wd, 'daily_report', None) and wd.daily_report.activities_done or '',
                'activities_planned': getattr(wd, 'daily_report', None) and wd.daily_report.activities_planned or '',
            }

        last_report = None
        if not workday_data:
            last = (
                Workday.objects
                .filter(employee=emp, status=Workday.STATUS_COMPLETED)
                .select_related('daily_report')
                .order_by('-end_time')
                .first()
            )
            if last:
                try:
                    last_report = {
                        'has_report': True,
                        'activities_done': last.daily_report.activities_done,
                        'activities_planned': last.daily_report.activities_planned,
                        'date': last.end_time.date(),
                    }
                except DailyReport.DoesNotExist:
                    pass

        return Response({**profile, 'workday': workday_data, 'last_report': last_report})


class PendingMessagesView(APIView):
    """Empleado obtiene sus mensajes pendientes de confirmar."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        messages = (
            ExecutiveMessage.objects
            .filter(recipient=employee, dismissed=False)
            .select_related('sender')
            .order_by('sent_at')
        )
        today = timezone.localdate()
        data = []
        for m in messages:
            is_own = bool(m.sender_id and m.sender_id == m.recipient_id)
            # La copia propia ("Enviado") no se confirma ni se borra; solo se ve el día en que se envió.
            if is_own and timezone.localtime(m.sent_at).date() != today:
                continue
            data.append({
                'id': m.id,
                'body': m.body,
                'sent_at': m.sent_at,
                'sender_name': m.sender.full_name if m.sender else 'skyBot',
                'day_closed': m.day_closed,
                'is_own': is_own,
                'acknowledged': m.acknowledged_at is not None,
            })
        return Response(data)


class WorkdayMonthlyView(APIView):
    """Devuelve horas trabajadas por día para el mes solicitado."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            employee = request.user.employee
        except Exception:
            return Response({})

        local_now = _local_now()
        try:
            year  = int(request.query_params.get('year',  local_now.year))
            month = int(request.query_params.get('month', local_now.month))
        except (TypeError, ValueError):
            return Response({'error': 'year y month deben ser enteros'}, status=400)

        first_dt, last_dt = _month_range(year, month)
        completed = Workday.objects.filter(
            employee=employee,
            status__in=[Workday.STATUS_COMPLETED, Workday.STATUS_INCOMPLETE],
            start_time__gte=first_dt,
            start_time__lte=last_dt,
        ).values('start_time', 'duration_minutes', 'auto_closed')

        days: dict = {}
        auto_closed_days: set = set()
        for w in completed:
            day = timezone.localtime(w['start_time']).day
            days[day] = days.get(day, 0) + (w['duration_minutes'] or 0) / 60
            if w['auto_closed']:
                auto_closed_days.add(day)

        # Incluir jornada activa si cae en el mes solicitado
        active_day = None
        try:
            active = Workday.objects.filter(employee=employee, status=Workday.STATUS_IN_PROGRESS).order_by('-start_time').first()
            if not active: raise Workday.DoesNotExist
            local_start = timezone.localtime(active.start_time)
            if local_start.year == year and local_start.month == month:
                active_day = local_start.day
                elapsed = int((local_now - active.start_time).total_seconds() // 60) / 60
                days[active_day] = days.get(active_day, 0) + elapsed
        except Workday.DoesNotExist:
            pass

        # Notas globales del mes
        notes_qs = CalendarNote.objects.filter(date__year=year, date__month=month)
        notes = {str(n.date.day): {'text': n.text, 'type': n.note_type} for n in notes_qs}

        # Ausencias del empleado en el mes
        first_date = _date(year, month, 1)
        last_date  = _date(year, month, _cal.monthrange(year, month)[1])
        leaves_qs = EmployeeLeave.objects.filter(
            employee=employee,
            start_date__lte=last_date,
            end_date__gte=first_date,
        )
        leave_days = {}
        for lv in leaves_qs:
            cur = max(lv.start_date, first_date)
            end = min(lv.end_date, last_date)
            while cur <= end:
                leave_days[str(cur.day)] = {'type': lv.leave_type, 'note': lv.note, 'id': lv.id}
                cur += timedelta(days=1)

        return Response({
            'year': year,
            'month': month,
            'active_day': active_day,
            'auto_closed_days': list(auto_closed_days),
            'days': {str(k): round(v, 2) for k, v in days.items()},
            'notes': notes,
            'leaves': leave_days,
        })


class EmployeeMonthlyView(APIView):
    """Ejecutivo ve el calendario de horas trabajadas de un empleado específico."""
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({})

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            employee = Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        local_now = _local_now()
        try:
            year  = int(request.query_params.get('year',  local_now.year))
            month = int(request.query_params.get('month', local_now.month))
        except (TypeError, ValueError):
            return Response({'error': 'year y month deben ser enteros'}, status=400)

        first_dt, last_dt = _month_range(year, month)
        completed = Workday.objects.filter(
            employee=employee,
            status__in=[Workday.STATUS_COMPLETED, Workday.STATUS_INCOMPLETE],
            start_time__gte=first_dt,
            start_time__lte=last_dt,
        ).values('start_time', 'duration_minutes', 'auto_closed')

        days: dict = {}
        auto_closed_days: set = set()
        for w in completed:
            day = timezone.localtime(w['start_time']).day
            days[day] = days.get(day, 0) + (w['duration_minutes'] or 0) / 60
            if w['auto_closed']:
                auto_closed_days.add(day)

        active_day = None
        try:
            active = Workday.objects.filter(employee=employee, status=Workday.STATUS_IN_PROGRESS).order_by('-start_time').first()
            if not active: raise Workday.DoesNotExist
            local_start = timezone.localtime(active.start_time)
            if local_start.year == year and local_start.month == month:
                active_day = local_start.day
                elapsed = int((local_now - active.start_time).total_seconds() // 60) / 60
                days[active_day] = days.get(active_day, 0) + elapsed
        except Workday.DoesNotExist:
            pass

        notes_qs = CalendarNote.objects.filter(date__year=year, date__month=month)
        notes = {str(n.date.day): {'text': n.text, 'type': n.note_type, 'id': n.id} for n in notes_qs}

        first_date = _date(year, month, 1)
        last_date  = _date(year, month, _cal.monthrange(year, month)[1])
        leaves_qs = EmployeeLeave.objects.filter(
            employee=employee,
            start_date__lte=last_date,
            end_date__gte=first_date,
        )
        leave_days = {}
        for lv in leaves_qs:
            cur = max(lv.start_date, first_date)
            end = min(lv.end_date, last_date)
            while cur <= end:
                leave_days[str(cur.day)] = {'type': lv.leave_type, 'note': lv.note, 'id': lv.id}
                cur += timedelta(days=1)

        return Response({
            'year': year,
            'month': month,
            'active_day': active_day,
            'auto_closed_days': list(auto_closed_days),
            'days': {str(k): round(v, 2) for k, v in days.items()},
            'notes': notes,
            'leaves': leave_days,
        })


class AcknowledgeMessageView(APIView):
    """Empleado confirma haber leído un mensaje."""
    permission_classes = [IsAuthenticated]

    def post(self, request, message_id):
        try:
            requester = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        view_as_id = request.query_params.get('view_as')
        if view_as_id and requester.is_executive:
            try:
                employee = Employee.objects.get(id=view_as_id, is_active=True)
            except Employee.DoesNotExist:
                return Response({'error': 'Empleado no encontrado'}, status=404)
        else:
            employee = requester

        try:
            message = ExecutiveMessage.objects.get(id=message_id, recipient=employee)
        except ExecutiveMessage.DoesNotExist:
            return Response({'error': 'Mensaje no encontrado'}, status=404)

        if message.acknowledged_at is None:
            message.acknowledged_at = timezone.now()
            message.save(update_fields=['acknowledged_at'])

        return Response({'ok': True})


class DismissMessageView(APIView):
    """El empleado descarta (oculta) un mensaje con la ✕. Requiere haberlo confirmado antes."""
    permission_classes = [IsAuthenticated]

    def post(self, request, message_id):
        try:
            requester = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        view_as_id = request.query_params.get('view_as')
        if view_as_id and requester.is_executive:
            try:
                employee = Employee.objects.get(id=view_as_id, is_active=True)
            except Employee.DoesNotExist:
                return Response({'error': 'Empleado no encontrado'}, status=404)
        else:
            employee = requester

        try:
            message = ExecutiveMessage.objects.get(id=message_id, recipient=employee)
        except ExecutiveMessage.DoesNotExist:
            return Response({'error': 'Mensaje no encontrado'}, status=404)

        # No se puede borrar sin confirmar (Listo) primero.
        if message.acknowledged_at is None:
            return Response({'error': 'Confirma el mensaje (Listo) antes de borrarlo'}, status=status.HTTP_400_BAD_REQUEST)

        if not message.dismissed:
            message.dismissed = True
            message.save(update_fields=['dismissed'])

        return Response({'ok': True})


class CalendarNotesView(APIView):
    """CRUD de notas globales de calendario (ejecutivos)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Lista notas de un mes: ?year=YYYY&month=MM"""
        now = timezone.now()
        try:
            year  = int(request.query_params.get('year',  now.year))
            month = int(request.query_params.get('month', now.month))
        except (TypeError, ValueError):
            return Response({'error': 'Parámetros inválidos'}, status=400)

        notes = CalendarNote.objects.filter(date__year=year, date__month=month)
        return Response([
            {'id': n.id, 'date': n.date, 'text': n.text, 'note_type': n.note_type}
            for n in notes
        ])

    def post(self, request):
        """Crea una nota. Solo ejecutivos."""
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        date_str  = request.data.get('date', '')
        text      = request.data.get('text', '').strip()
        note_type = request.data.get('note_type', CalendarNote.TYPE_FERIADO)

        if not date_str or not text:
            return Response({'error': 'date y text son requeridos'}, status=400)
        try:
            from datetime import date as date_type
            import datetime
            date_obj = datetime.date.fromisoformat(date_str)
        except ValueError:
            return Response({'error': 'Formato de fecha inválido (YYYY-MM-DD)'}, status=400)

        note, created = CalendarNote.objects.update_or_create(
            date=date_obj,
            defaults={'text': text, 'note_type': note_type, 'created_by': executive},
        )
        return Response(
            {'id': note.id, 'date': note.date, 'text': note.text, 'note_type': note.note_type},
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class CalendarNoteDetailView(APIView):
    """Elimina una nota de calendario. Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, note_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)
        try:
            note = CalendarNote.objects.get(id=note_id)
        except CalendarNote.DoesNotExist:
            return Response({'error': 'Nota no encontrada'}, status=404)
        note.delete()
        return Response({'ok': True})


class EmployeeLeavesView(APIView):
    """Lista y crea ausencias de un empleado. Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        now = timezone.now()
        try:
            year  = int(request.query_params.get('year',  now.year))
            month = int(request.query_params.get('month', now.month))
        except (TypeError, ValueError):
            return Response({'error': 'Parámetros inválidos'}, status=400)

        leaves = EmployeeLeave.objects.filter(
            employee_id=employee_id,
            start_date__year=year,
            start_date__month=month,
        )
        return Response([
            {
                'id': l.id,
                'start_date': l.start_date,
                'end_date': l.end_date,
                'leave_type': l.leave_type,
                'note': l.note,
            }
            for l in leaves
        ])

    def post(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            emp = Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        start_str  = request.data.get('start_date', '')
        end_str    = request.data.get('end_date', '') or start_str
        leave_type = request.data.get('leave_type', '')
        note       = request.data.get('note', '').strip()

        if not start_str or not leave_type:
            return Response({'error': 'start_date y leave_type son requeridos'}, status=400)

        try:
            import datetime
            start = datetime.date.fromisoformat(start_str)
            end   = datetime.date.fromisoformat(end_str)
            if end < start:
                end = start
        except ValueError:
            return Response({'error': 'Formato de fecha inválido (YYYY-MM-DD)'}, status=400)

        leave = EmployeeLeave.objects.create(
            employee=emp,
            start_date=start,
            end_date=end,
            leave_type=leave_type,
            note=note,
            created_by=executive,
        )
        return Response(
            {'id': leave.id, 'start_date': leave.start_date, 'end_date': leave.end_date,
             'leave_type': leave.leave_type, 'note': leave.note},
            status=status.HTTP_201_CREATED,
        )


class EmployeeLeaveDetailView(APIView):
    """Elimina una ausencia. Solo ejecutivos."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, employee_id, leave_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)
        try:
            leave = EmployeeLeave.objects.get(id=leave_id, employee_id=employee_id)
        except EmployeeLeave.DoesNotExist:
            return Response({'error': 'Ausencia no encontrada'}, status=404)
        leave.delete()
        return Response({'ok': True})


_MONTH_NAMES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
_DAY_NAMES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']


def reporte_view(request):
    """Sirve el shell HTML del reporte. Los datos se cargan desde /api/reporte/ via JS."""
    local_now = _local_now()
    years_range = list(range(local_now.year - 2, local_now.year + 1))
    return render(request, 'workdays/reporte.html', {
        'current_year': local_now.year,
        'current_month': local_now.month,
        'months': list(enumerate(_MONTH_NAMES, 1)),
        'years': years_range,
    })


class ReporteAPIView(APIView):
    """Datos del reporte de asistencia. Ejecutivos o usuarios con can_view_report."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (executive.is_executive or executive.can_view_report):
            return Response({'error': 'Acceso no autorizado'}, status=403)

        local_now = _local_now()
        filter_mode = request.query_params.get('mode', 'month')

        if filter_mode == 'range':
            from_str = request.query_params.get('from', '')
            to_str   = request.query_params.get('to', '')
            try:
                from_date = _date.fromisoformat(from_str)
                to_date   = _date.fromisoformat(to_str)
                if to_date < from_date:
                    to_date = from_date
            except (ValueError, TypeError):
                filter_mode = 'month'

        if filter_mode == 'month':
            try:
                sel_year  = int(request.query_params.get('year',  local_now.year))
                sel_month = int(request.query_params.get('month', local_now.month))
                if not (1 <= sel_month <= 12):
                    raise ValueError
            except (ValueError, TypeError):
                sel_year, sel_month = local_now.year, local_now.month
            from_date = _date(sel_year, sel_month, 1)
            to_date   = _date(sel_year, sel_month, _cal.monthrange(sel_year, sel_month)[1])
            label = f'{_MONTH_NAMES[sel_month - 1]} {sel_year}'
        else:
            label = f'{from_date.strftime("%d/%m/%Y")} — {to_date.strftime("%d/%m/%Y")}'

        today = local_now.date()
        if to_date > today:
            to_date = today

        first_dt = timezone.make_aware(_datetime.combine(from_date, _time(0, 0, 0)))
        last_dt  = timezone.make_aware(_datetime.combine(to_date,   _time(23, 59, 59)))

        from employees.models import Employee as _Employee
        employees = list(_Employee.objects.filter(is_active=True, is_executive=False).order_by('full_name'))

        # Workday lookup: (emp_id, date) -> workday
        wd_lookup = {}
        for wd in Workday.objects.filter(
            employee__is_active=True, employee__is_executive=False,
            status__in=[Workday.STATUS_COMPLETED, Workday.STATUS_INCOMPLETE],
            start_time__gte=first_dt, start_time__lte=last_dt,
        ).select_related('employee'):
            local_start = timezone.localtime(wd.start_time)
            wd_lookup[(wd.employee_id, local_start.date())] = wd

        # Leaves lookup: (emp_id, date) -> [texts]
        leave_type_labels = dict(EmployeeLeave.TYPE_CHOICES)
        leaves_lookup = {}
        for lv in EmployeeLeave.objects.filter(
            employee__is_active=True, employee__is_executive=False,
            start_date__lte=to_date, end_date__gte=from_date,
        ):
            d = lv.start_date
            while d <= lv.end_date:
                if from_date <= d <= to_date:
                    txt = leave_type_labels.get(lv.leave_type, lv.leave_type)
                    if lv.note:
                        txt += f' ({lv.note})'
                    leaves_lookup.setdefault((lv.employee_id, d), []).append(txt)
                d += timedelta(days=1)

        note_type_labels = dict(CalendarNote.TYPE_CHOICES)
        notes_lookup = {}
        for note in CalendarNote.objects.filter(date__gte=from_date, date__lte=to_date):
            lbl = note_type_labels.get(note.note_type, note.note_type)
            notes_lookup.setdefault(note.date, []).append(f'{lbl}: {note.text}')

        all_dates = []
        d = from_date
        while d <= to_date:
            all_dates.append(d)
            d += timedelta(days=1)

        rows = []
        for emp_counter, emp in enumerate(employees, start=1):
            for day_num, day in enumerate(all_dates, start=1):
                wd = wd_lookup.get((emp.id, day))
                if wd:
                    local_start = timezone.localtime(wd.start_time)
                    local_end   = timezone.localtime(wd.end_time) if wd.end_time else None
                    ref_mins    = emp.hora_entrada.hour * 60 + emp.hora_entrada.minute
                    start_mins  = local_start.hour * 60 + local_start.minute
                    atraso      = max(0, start_mins - ref_mins)
                    neto        = max(0, (wd.duration_minutes or 0) - 60)
                    hora_ingreso    = local_start.strftime('%H:%M')
                    hora_salida     = local_end.strftime('%H:%M') if local_end else '—'
                    horas_trabajadas = f'{neto // 60:02d}:{neto % 60:02d}'
                    atraso_minutos  = atraso
                else:
                    hora_ingreso = hora_salida = horas_trabajadas = ''
                    atraso_minutos = None

                rows.append({
                    'emp_num':   emp_counter,
                    'day_num':   day_num,
                    'nombre':    emp.full_name,
                    'cargo':     emp.cargo,
                    'haber_basico': str(emp.haber_basico) if emp.haber_basico else None,
                    'fecha':     day.strftime('%d-%m-%Y'),
                    'dia':       _DAY_NAMES[day.weekday()],
                    'hora_ingreso':    hora_ingreso,
                    'hora_salida':     hora_salida,
                    'horas_trabajadas': horas_trabajadas,
                    'atraso_minutos':  atraso_minutos,
                    'no_cerro':        bool(wd and wd.auto_closed),
                    'comentario_leaves': leaves_lookup.get((emp.id, day), []),
                    'comentario_notes':  notes_lookup.get(day, []),
                    'is_weekend':      day.weekday() >= 5,
                    'is_first_of_employee': day_num == 1,
                })

        return Response({'rows': rows, 'total': len(rows), 'label': label})


def _parse_period(request):
    """Parsea mode=month|range de los query params y devuelve el rango.

    Returns: (first_dt, last_dt, from_date, to_date, label)
    """
    local_now = _local_now()
    filter_mode = request.query_params.get('mode', 'month')

    from_date = to_date = None
    if filter_mode == 'range':
        try:
            from_date = _date.fromisoformat(request.query_params.get('from', ''))
            to_date   = _date.fromisoformat(request.query_params.get('to', ''))
            if to_date < from_date:
                to_date = from_date
        except (ValueError, TypeError):
            filter_mode = 'month'

    if filter_mode == 'month':
        try:
            sel_year  = int(request.query_params.get('year',  local_now.year))
            sel_month = int(request.query_params.get('month', local_now.month))
            if not (1 <= sel_month <= 12):
                raise ValueError
        except (ValueError, TypeError):
            sel_year, sel_month = local_now.year, local_now.month
        from_date = _date(sel_year, sel_month, 1)
        to_date   = _date(sel_year, sel_month, _cal.monthrange(sel_year, sel_month)[1])
        label = f'{_MONTH_NAMES[sel_month - 1]} {sel_year}'
    else:
        label = f'{from_date.strftime("%d/%m/%Y")} — {to_date.strftime("%d/%m/%Y")}'

    today = local_now.date()
    if to_date > today:
        to_date = today

    first_dt = timezone.make_aware(_datetime.combine(from_date, _time(0, 0, 0)))
    last_dt  = timezone.make_aware(_datetime.combine(to_date,   _time(23, 59, 59)))
    return first_dt, last_dt, from_date, to_date, label


def _tag_breakdown(qs, kind, top=12, split_by_sede=False, total_qs=None):
    """Top-N tags de un `kind` sobre el queryset de ActivityItem ya filtrado,
    agrupando alias en su tag canónico (root). Devuelve [{id, name, color, count}].
    Con split_by_sede=True agrega `segments` (desglose por sede del empleado).
    Con total_qs agrega `total` (conteo global, sin el filtro de empleado) para comparar."""
    from collections import defaultdict
    from django.db.models import Count
    rows = (qs.filter(tags__kind=kind, tags__ignored=False)
            .values('tags__id', 'tags__name', 'tags__color',
                    'tags__canonical_id', 'tags__canonical__name', 'tags__canonical__color')
            .annotate(n=Count('id')))
    agg = {}
    for r in rows:
        cid = r['tags__canonical_id'] or r['tags__id']
        if cid is None:
            continue
        name = r['tags__canonical__name'] or r['tags__name']
        color = (r['tags__canonical__color'] if r['tags__canonical_id'] else r['tags__color']) or ''
        bucket = agg.setdefault(cid, {'id': cid, 'name': name, 'color': color, 'count': 0})
        bucket['count'] += r['n']

    if split_by_sede:
        from .models import ActivityTag, SEDE_NAMES
        sede_color = dict(ActivityTag.objects.filter(
            kind=ActivityTag.KIND_LOCATION, canonical__isnull=True).values_list('name', 'color'))
        seg = defaultdict(lambda: defaultdict(int))   # cid -> sede -> count
        for r in (qs.filter(tags__kind=kind, tags__ignored=False)
                  .values('tags__id', 'tags__canonical_id', 'report__workday__employee__ciudad')
                  .annotate(n=Count('id'))):
            cid = r['tags__canonical_id'] or r['tags__id']
            if cid is None:
                continue
            sede = SEDE_NAMES.get((r['report__workday__employee__ciudad'] or '').strip().upper(), '—')
            seg[cid][sede] += r['n']
        for b in agg.values():
            parts = sorted(seg.get(b['id'], {}).items(), key=lambda x: (-x[1], x[0]))
            b['segments'] = [{'name': s, 'count': c, 'color': sede_color.get(s, '')} for s, c in parts]

    if total_qs is not None:
        totals = defaultdict(int)
        for r in (total_qs.filter(tags__kind=kind, tags__ignored=False)
                  .values('tags__id', 'tags__canonical_id').annotate(n=Count('id'))):
            cid = r['tags__canonical_id'] or r['tags__id']
            if cid is not None:
                totals[cid] += r['n']
        for b in agg.values():
            b['total'] = totals.get(b['id'], b['count'])

    out = sorted(agg.values(), key=lambda x: (-x['count'], x['name'].lower()))
    return out[:top]


def estadisticas_view(request):
    """Sirve el shell HTML de estadísticas. Datos vía /api/estadisticas/."""
    local_now = _local_now()
    years_range = list(range(local_now.year - 2, local_now.year + 1))
    return render(request, 'workdays/estadisticas.html', {
        'current_year': local_now.year,
        'current_month': local_now.month,
        'months': list(enumerate(_MONTH_NAMES, 1)),
        'years': years_range,
    })


class EstadisticasAPIView(APIView):
    """Estadísticas de actividades clasificadas. Solo ejecutivos.

    Excluye ejecutivos y empleados con rol "otro" (almacén): las estadísticas
    son sobre Supervisores vs Project Managers.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from collections import Counter, defaultdict
        from django.db.models.functions import TruncMonth, TruncWeek
        from .models import ActivityItem, ActivityCategory

        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (executive.is_executive or executive.can_view_stats):
            return Response({'error': 'Acceso no autorizado'}, status=403)

        mode = request.query_params.get('mode', 'month')
        if mode == 'all':
            first_dt = last_dt = from_date = to_date = None
            label = 'Todos'
            period = ''  # se completa con el rango real de los datos más abajo
        elif mode == 'last':
            # El último día con datos se resuelve más abajo, una vez aplicados los filtros.
            first_dt = last_dt = from_date = to_date = None
            label = 'Último día'
            period = ''
        else:
            first_dt, last_dt, from_date, to_date, label = _parse_period(request)
            period = f'{from_date.strftime("%d/%m/%Y")} — {to_date.strftime("%d/%m/%Y")}'
        kind = request.query_params.get('kind', ActivityItem.KIND_DONE)
        if kind not in (ActivityItem.KIND_DONE, ActivityItem.KIND_PLANNED):
            kind = ActivityItem.KIND_DONE
        role_filter = request.query_params.get('role') or ''
        emp_filter  = request.query_params.get('employee') or ''

        # Empleados elegibles: activos, no ejecutivos, rol supervisor/PM.
        emp_meta = {}  # id -> (name, role)
        for emp in Employee.objects.filter(is_active=True, is_executive=False):
            if emp.role == Employee.ROLE_OTRO:
                continue
            emp_meta[emp.id] = (emp.full_name, emp.role)

        cats = list(ActivityCategory.objects.order_by('order', 'label'))
        cat_codes  = [c.code for c in cats]
        cat_labels = {c.code: c.label for c in cats}
        cat_colors = {c.code: c.color for c in cats}

        def _base_qs(dt_from, dt_to, apply_emp=True):
            qs = ActivityItem.objects.filter(
                kind=kind,
                report__workday__employee_id__in=emp_meta.keys(),
            )
            if dt_from is not None:
                qs = qs.filter(report__workday__start_time__gte=dt_from)
            if dt_to is not None:
                qs = qs.filter(report__workday__start_time__lte=dt_to)
            if apply_emp and emp_filter:
                qs = qs.filter(report__workday__employee_id=emp_filter)
            if role_filter:
                ids = [eid for eid, (_, r) in emp_meta.items() if r == role_filter]
                qs = qs.filter(report__workday__employee_id__in=ids)
            return qs

        # En modo "Todo": etiqueta con el rango real (reporte más antiguo → más reciente).
        if mode == 'all':
            agg = _base_qs(None, None).aggregate(
                mn=models.Min('report__workday__start_time'),
                mx=models.Max('report__workday__start_time'))
            if agg['mn'] and agg['mx']:
                d0 = timezone.localtime(agg['mn']).strftime('%d/%m/%Y')
                d1 = timezone.localtime(agg['mx']).strftime('%d/%m/%Y')
                period = f'{d0} — {d1}'  # el rango va solo en "Período:"; label queda 'Todos'

        # En modo "Último": acota al día (local) del reporte más reciente con datos.
        elif mode == 'last':
            mx = _base_qs(None, None).aggregate(
                mx=models.Max('report__workday__start_time'))['mx']
            if mx:
                day = timezone.localtime(mx).date()
                from_date = to_date = day
                first_dt = timezone.make_aware(_datetime.combine(day, _time(0, 0, 0)))
                last_dt  = timezone.make_aware(_datetime.combine(day, _time(23, 59, 59)))
                period = day.strftime('%d/%m/%Y')  # el día va solo en "Período:"; label queda 'Último día'

        rows = list(_base_qs(first_dt, last_dt).values_list(
            'category', 'report__workday__employee_id', 'report_id',
            'report__workday__start_time'))

        total_items = len(rows)
        cat_counter = Counter()
        emp_counter = defaultdict(lambda: defaultdict(int))    # emp_id -> cat -> n
        emp_reports = defaultdict(set)                          # emp_id -> {report_id}
        emp_last = {}                                          # emp_id -> max start_time
        report_ids = set()

        for cat, emp_id, report_id, st in rows:
            cat_counter[cat] += 1
            emp_counter[emp_id][cat] += 1
            emp_reports[emp_id].add(report_id)
            if emp_id not in emp_last or st > emp_last[emp_id]:
                emp_last[emp_id] = st
            report_ids.add(report_id)

        # Referencia por rol (ignora el filtro de empleado): se agrega por empleado para
        # poder calcular promedios por empleado y compararlos con el nombre seleccionado.
        ref_emp_cat = defaultdict(lambda: defaultdict(int))    # emp_id -> cat -> n
        ref_emp_reports = defaultdict(set)                     # emp_id -> {report_id}
        for cat, emp_id, rep_id in _base_qs(first_dt, last_dt, apply_emp=False).values_list(
                'category', 'report__workday__employee_id', 'report_id'):
            ref_emp_cat[emp_id][cat] += 1
            ref_emp_reports[emp_id].add(rep_id)

        role_counter = defaultdict(lambda: defaultdict(int))   # role -> cat -> n
        role_total = defaultdict(int)
        role_emps = defaultdict(set)                           # role -> {emp_id}
        role_report_days = defaultdict(int)                    # role -> suma de días c/reporte
        for emp_id, ctr in ref_emp_cat.items():
            _, role = emp_meta[emp_id]
            role_emps[role].add(emp_id)
            role_report_days[role] += len(ref_emp_reports[emp_id])
            for c, n in ctr.items():
                role_counter[role][c] += n
                role_total[role] += n

        categories = [
            {'code': code, 'label': cat_labels[code], 'count': cat_counter.get(code, 0),
             'pct': round(100 * cat_counter.get(code, 0) / total_items, 1) if total_items else 0}
            for code in cat_codes
        ]

        by_role = []
        for role in (Employee.ROLE_SUPERVISOR, Employee.ROLE_PROJECT_MANAGER):
            if not role_total.get(role):
                continue
            n = len(role_emps[role]) or 1
            rdays = role_report_days[role]
            by_role.append({
                'role': role,
                'label': Employee.ROLE_LABELS[role],
                'total': role_total[role],
                'categories': {code: role_counter[role].get(code, 0) for code in cat_codes},
                # Promedios por empleado del rol (referencia para tablas/top categorías).
                'n_emps': len(role_emps[role]),
                'avg_total': round(role_total[role] / n, 1),
                'avg_reports': round(rdays / n, 1),
                'avg_perday': round(role_total[role] / rdays, 1) if rdays else 0,
                'avg_categories': {code: round(role_counter[role].get(code, 0) / n, 1) for code in cat_codes},
            })

        employees = []
        for emp_id, counter in emp_counter.items():
            name, role = emp_meta[emp_id]
            tot = sum(counter.values())
            top = max(counter, key=counter.get)
            last = emp_last.get(emp_id)
            employees.append({
                'id': emp_id, 'name': name, 'role': role, 'total': tot,
                'reports': len(emp_reports[emp_id]),
                'last_date': timezone.localtime(last).strftime('%d/%m/%Y') if last else '',
                'top_category': top,
                'categories': {code: counter.get(code, 0) for code in cat_codes},
            })
        employees.sort(key=lambda e: (-e['total'], e['name']))

        # Evolución temporal: mensual (últimos 6 meses) o semanal (últimas 12 semanas).
        # En "Todo": todos los periodos con datos.
        gran = request.query_params.get('granularity', 'month')
        if gran not in ('month', 'week'):
            gran = 'month'

        if mode == 'all':
            ev_first = None
        elif gran == 'week':
            ev_start = from_date - timedelta(days=from_date.weekday(), weeks=11)
            ev_first = timezone.make_aware(_datetime.combine(ev_start, _time(0, 0, 0)))
        else:
            ev_start = from_date.replace(day=1)
            for _ in range(5):
                ev_start = (ev_start - timedelta(days=1)).replace(day=1)
            ev_first = timezone.make_aware(_datetime.combine(ev_start, _time(0, 0, 0)))

        Trunc = TruncWeek if gran == 'week' else TruncMonth
        evo_rows = _base_qs(ev_first, last_dt).annotate(
            p=Trunc('report__workday__start_time')
        ).values_list('p', 'category')
        bucket = defaultdict(lambda: defaultdict(int))
        for p, cat in evo_rows:
            lp = timezone.localtime(p) if timezone.is_aware(p) else p
            key = lp.strftime('%Y-%m-%d') if gran == 'week' else lp.strftime('%Y-%m')
            bucket[key][cat] += 1
        monthly = []
        for key in sorted(bucket):
            cats = bucket[key]
            if gran == 'week':
                lbl = _date.fromisoformat(key).strftime('%d/%m')
            else:
                y, mm = key.split('-')
                lbl = f'{_MONTH_NAMES[int(mm) - 1][:3]} {y[2:]}'
            monthly.append({
                'month': key,
                'label': lbl,
                'total': sum(cats.values()),
                'categories': {code: cats.get(code, 0) for code in cat_codes},
            })

        # Empleado seleccionado (si hay filtro): su composición propia, para
        # graficarla contra el promedio de cada rol como referencia.
        selected_employee = None
        if emp_filter:
            try:
                sel_id = int(emp_filter)
            except (TypeError, ValueError):
                sel_id = None
            if sel_id in emp_meta:
                name, role = emp_meta[sel_id]
                ctr = emp_counter.get(sel_id, {})
                selected_employee = {
                    'id': sel_id,
                    'name': name,
                    'role': role,
                    'total': sum(ctr.values()),
                    'categories': {code: ctr.get(code, 0) for code in cat_codes},
                }

        # Con un empleado seleccionado, mostramos su conteo + el TOTAL de todos
        # (queryset sin el filtro de empleado) para comparar.
        tot_qs = _base_qs(first_dt, last_dt, apply_emp=False) if emp_filter else None
        return Response({
            'label': label,
            'period': period,
            'kind': kind,
            'granularity': gran,
            'total_items': total_items,
            'total_reports': len(report_ids),
            'category_labels': cat_labels,
            'category_colors': cat_colors,
            'categories': categories,
            'by_role': by_role,
            'selected_employee': selected_employee,
            'monthly': monthly,
            'employees': employees,
            'by_project':     _tag_breakdown(_base_qs(first_dt, last_dt), 'project', split_by_sede=True, total_qs=tot_qs),
            'by_location':    _tag_breakdown(_base_qs(first_dt, last_dt), 'location'),
            'by_deliverable': _tag_breakdown(_base_qs(first_dt, last_dt), 'deliverable', split_by_sede=True, total_qs=tot_qs),
        })


class MisEstadisticasView(APIView):
    """Estadísticas de actividades del propio empleado logueado.

    Igual estructura que EstadisticasAPIView pero acotado al usuario actual y SIN
    requerir is_executive — cada empleado ve solo sus propios datos.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from collections import Counter, defaultdict
        from django.db.models.functions import TruncMonth, TruncWeek
        from .models import ActivityItem, ActivityCategory

        try:
            employee = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        # Ejecutivo viendo el dashboard de un empleado (iframe view_as).
        view_as_id = request.query_params.get('view_as')
        if view_as_id and getattr(employee, 'is_executive', False):
            try:
                employee = Employee.objects.get(id=view_as_id, is_active=True)
            except Employee.DoesNotExist:
                pass

        mode = request.query_params.get('mode', 'all')
        if mode in ('all', 'last'):
            first_dt = last_dt = from_date = to_date = None
            label = 'Todos' if mode == 'all' else 'Último día'
            period = ''
        else:
            first_dt, last_dt, from_date, to_date, label = _parse_period(request)
            period = f'{from_date.strftime("%d/%m/%Y")} — {to_date.strftime("%d/%m/%Y")}'

        kind = request.query_params.get('kind', ActivityItem.KIND_DONE)
        if kind not in (ActivityItem.KIND_DONE, ActivityItem.KIND_PLANNED):
            kind = ActivityItem.KIND_DONE

        cats = list(ActivityCategory.objects.order_by('order', 'label'))
        cat_codes  = [c.code for c in cats]
        cat_labels = {c.code: c.label for c in cats}
        cat_colors = {c.code: c.color for c in cats}

        def _base_qs(dt_from, dt_to):
            qs = ActivityItem.objects.filter(kind=kind, report__workday__employee_id=employee.id)
            if dt_from is not None:
                qs = qs.filter(report__workday__start_time__gte=dt_from)
            if dt_to is not None:
                qs = qs.filter(report__workday__start_time__lte=dt_to)
            return qs

        if mode == 'all':
            agg = _base_qs(None, None).aggregate(
                mn=models.Min('report__workday__start_time'),
                mx=models.Max('report__workday__start_time'))
            if agg['mn'] and agg['mx']:
                d0 = timezone.localtime(agg['mn']).strftime('%d/%m/%Y')
                d1 = timezone.localtime(agg['mx']).strftime('%d/%m/%Y')
                period = f'{d0} — {d1}'
        elif mode == 'last':
            mx = _base_qs(None, None).aggregate(mx=models.Max('report__workday__start_time'))['mx']
            if mx:
                day = timezone.localtime(mx).date()
                from_date = to_date = day
                first_dt = timezone.make_aware(_datetime.combine(day, _time(0, 0, 0)))
                last_dt  = timezone.make_aware(_datetime.combine(day, _time(23, 59, 59)))
                period = day.strftime('%d/%m/%Y')

        rows = list(_base_qs(first_dt, last_dt).values_list(
            'category', 'report_id', 'report__workday__start_time'))
        total_items = len(rows)
        cat_counter = Counter()
        report_ids = set()
        last_dt_seen = None
        for cat, rep_id, st in rows:
            cat_counter[cat] += 1
            report_ids.add(rep_id)
            if last_dt_seen is None or st > last_dt_seen:
                last_dt_seen = st

        categories = [
            {'code': code, 'label': cat_labels[code], 'count': cat_counter.get(code, 0),
             'pct': round(100 * cat_counter.get(code, 0) / total_items, 1) if total_items else 0}
            for code in cat_codes
        ]

        # Evolución temporal (semanal: últimas 12 semanas; mensual: últimos 6 meses).
        gran = request.query_params.get('granularity', 'week')
        if gran not in ('month', 'week'):
            gran = 'week'
        if mode == 'all':
            ev_first = None
        elif gran == 'week':
            ev_start = from_date - timedelta(days=from_date.weekday(), weeks=11)
            ev_first = timezone.make_aware(_datetime.combine(ev_start, _time(0, 0, 0)))
        else:
            ev_start = from_date.replace(day=1)
            for _ in range(5):
                ev_start = (ev_start - timedelta(days=1)).replace(day=1)
            ev_first = timezone.make_aware(_datetime.combine(ev_start, _time(0, 0, 0)))

        Trunc = TruncWeek if gran == 'week' else TruncMonth
        evo_rows = _base_qs(ev_first, last_dt).annotate(
            p=Trunc('report__workday__start_time')).values_list('p', 'category')
        bucket = defaultdict(lambda: defaultdict(int))
        for p, cat in evo_rows:
            lp = timezone.localtime(p) if timezone.is_aware(p) else p
            key = lp.strftime('%Y-%m-%d') if gran == 'week' else lp.strftime('%Y-%m')
            bucket[key][cat] += 1
        monthly = []
        for key in sorted(bucket):
            cats = bucket[key]
            if gran == 'week':
                lbl = _date.fromisoformat(key).strftime('%d/%m')
            else:
                y, mm = key.split('-')
                lbl = f'{_MONTH_NAMES[int(mm) - 1][:3]} {y[2:]}'
            monthly.append({
                'month': key, 'label': lbl, 'total': sum(cats.values()),
                'categories': {code: cats.get(code, 0) for code in cat_codes},
            })

        # Promedios por rol (referencia): sobre todos los empleados elegibles del
        # período, para comparar al empleado con el promedio de Supervisores y PMs.
        elig = {}  # emp_id -> role
        for e in Employee.objects.filter(is_active=True, is_executive=False):
            if e.role == Employee.ROLE_OTRO:
                continue
            elig[e.id] = e.role
        ref_qs = ActivityItem.objects.filter(kind=kind, report__workday__employee_id__in=elig.keys())
        if first_dt is not None:
            ref_qs = ref_qs.filter(report__workday__start_time__gte=first_dt)
        if last_dt is not None:
            ref_qs = ref_qs.filter(report__workday__start_time__lte=last_dt)
        ref_emp_cat = defaultdict(lambda: defaultdict(int))
        ref_emp_reports = defaultdict(set)
        for cat, eid, rid in ref_qs.values_list('category', 'report__workday__employee_id', 'report_id'):
            ref_emp_cat[eid][cat] += 1
            ref_emp_reports[eid].add(rid)
        role_counter = defaultdict(lambda: defaultdict(int))
        role_total = defaultdict(int)
        role_emps = defaultdict(set)
        role_rdays = defaultdict(int)
        for eid, ctr in ref_emp_cat.items():
            role = elig[eid]
            role_emps[role].add(eid)
            role_rdays[role] += len(ref_emp_reports[eid])
            for c, n in ctr.items():
                role_counter[role][c] += n
                role_total[role] += n
        by_role = []
        for role in (Employee.ROLE_SUPERVISOR, Employee.ROLE_PROJECT_MANAGER):
            if not role_total.get(role):
                continue
            n = len(role_emps[role]) or 1
            rd = role_rdays[role]
            by_role.append({
                'role': role,
                'label': Employee.ROLE_LABELS[role],
                'total': role_total[role],
                'categories': {code: role_counter[role].get(code, 0) for code in cat_codes},
                'n_emps': len(role_emps[role]),
                'avg_total': round(role_total[role] / n, 1),
                'avg_reports': round(rd / n, 1),
                'avg_perday': round(role_total[role] / rd, 1) if rd else 0,
                'avg_categories': {code: round(role_counter[role].get(code, 0) / n, 1) for code in cat_codes},
            })

        # Mismo shape que /estadisticas/ con un empleado seleccionado: el empleado
        # logueado es siempre el "seleccionado", para reusar el render de esa página.
        self_cats = {code: cat_counter.get(code, 0) for code in cat_codes}
        selected_employee = {
            'id': employee.id, 'name': employee.full_name, 'role': employee.role,
            'total': total_items, 'categories': self_cats,
        }
        employees = [{
            'id': employee.id, 'name': employee.full_name, 'role': employee.role,
            'total': total_items, 'reports': len(report_ids),
            'last_date': timezone.localtime(last_dt_seen).strftime('%d/%m/%Y') if last_dt_seen else '',
            'top_category': max(self_cats, key=self_cats.get) if total_items else '',
            'categories': self_cats,
        }]

        return Response({
            'label': label, 'period': period, 'kind': kind, 'granularity': gran,
            'total_items': total_items, 'total_reports': len(report_ids),
            'category_labels': cat_labels, 'category_colors': cat_colors,
            'categories': categories, 'monthly': monthly,
            'by_role': by_role, 'selected_employee': selected_employee, 'employees': employees,
            'by_project':     _tag_breakdown(_base_qs(first_dt, last_dt), 'project', split_by_sede=True, total_qs=ref_qs),
            'by_location':    _tag_breakdown(_base_qs(first_dt, last_dt), 'location', total_qs=ref_qs),
            'by_deliverable': _tag_breakdown(_base_qs(first_dt, last_dt), 'deliverable', split_by_sede=True, total_qs=ref_qs),
        })


def _xls_sheet_name(nombre):
    parts = nombre.strip().split()
    inicial = (parts[0][0].upper() + '.') if parts else ''
    apellido = parts[1] if len(parts) > 1 else ''
    name = f'{inicial} {apellido}'.strip()[:31]
    for ch in r'\/*?:[]':
        name = name.replace(ch, '')
    return name or 'Empleado'


class ReporteExportView(APIView):
    """Descarga el reporte de asistencia como .xlsx usando la plantilla corporativa."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not (executive.is_executive or executive.can_view_report):
            return Response({'error': 'Acceso no autorizado'}, status=403)

        local_now = _local_now()
        filter_mode = request.query_params.get('mode', 'month')

        if filter_mode == 'range':
            try:
                from_date = _date.fromisoformat(request.query_params.get('from', ''))
                to_date   = _date.fromisoformat(request.query_params.get('to', ''))
                if to_date < from_date:
                    to_date = from_date
            except (ValueError, TypeError):
                filter_mode = 'month'

        if filter_mode == 'month':
            try:
                sel_year  = int(request.query_params.get('year',  local_now.year))
                sel_month = int(request.query_params.get('month', local_now.month))
                if not (1 <= sel_month <= 12):
                    raise ValueError
            except (ValueError, TypeError):
                sel_year, sel_month = local_now.year, local_now.month
            from_date = _date(sel_year, sel_month, 1)
            to_date   = _date(sel_year, sel_month, _cal.monthrange(sel_year, sel_month)[1])
            label = f'{_MONTH_NAMES[sel_month - 1]}-{sel_year}'
        else:
            label = f'{from_date.strftime("%d-%m-%Y")}__{to_date.strftime("%d-%m-%Y")}'

        today = local_now.date()
        if to_date > today:
            to_date = today

        first_dt = timezone.make_aware(_datetime.combine(from_date, _time(0, 0, 0)))
        last_dt  = timezone.make_aware(_datetime.combine(to_date,   _time(23, 59, 59)))

        from employees.models import Employee as _EmpXls
        xls_employees = list(_EmpXls.objects.filter(is_active=True, is_executive=False).order_by('full_name'))

        # Workday lookup
        xls_wd_lookup = {}
        for wd in Workday.objects.filter(
            employee__is_active=True, employee__is_executive=False,
            status__in=[Workday.STATUS_COMPLETED, Workday.STATUS_INCOMPLETE],
            start_time__gte=first_dt, start_time__lte=last_dt,
        ).select_related('employee'):
            ls = timezone.localtime(wd.start_time)
            xls_wd_lookup[(wd.employee_id, ls.date())] = wd

        # Per-date lookups for comments
        xls_leave_labels = dict(EmployeeLeave.TYPE_CHOICES)
        xls_leaves_lookup = {}
        emp_ids_xls = {e.id for e in xls_employees}
        for lv in EmployeeLeave.objects.filter(
            employee_id__in=emp_ids_xls,
            start_date__lte=to_date, end_date__gte=from_date,
        ):
            d = lv.start_date
            while d <= lv.end_date:
                if from_date <= d <= to_date:
                    txt = xls_leave_labels.get(lv.leave_type, lv.leave_type)
                    if lv.note:
                        txt += f' ({lv.note})'
                    xls_leaves_lookup.setdefault((lv.employee_id, d), []).append(txt)
                d += timedelta(days=1)

        xls_note_labels = dict(CalendarNote.TYPE_CHOICES)
        xls_notes_lookup = {}
        for note in CalendarNote.objects.filter(date__gte=from_date, date__lte=to_date):
            lbl = xls_note_labels.get(note.note_type, note.note_type)
            xls_notes_lookup.setdefault(note.date, []).append(f'{lbl}: {note.text}')

        xls_all_dates = []
        d = from_date
        while d <= to_date:
            xls_all_dates.append(d)
            d += timedelta(days=1)

        # Agrupar por empleado con todos los días
        groups = {}
        order = []
        for emp_counter, emp in enumerate(xls_employees, start=1):
            groups[emp_counter] = {
                'emp_num': emp_counter,
                'nombre':  emp.full_name,
                'cargo':   emp.cargo,
                'haber_basico': float(emp.haber_basico) if emp.haber_basico else '',
                'rows': [],
            }
            order.append(emp_counter)
            for day in xls_all_dates:
                wd = xls_wd_lookup.get((emp.id, day))
                if wd:
                    ls  = timezone.localtime(wd.start_time)
                    le  = timezone.localtime(wd.end_time) if wd.end_time else None
                    ref = emp.hora_entrada.hour * 60 + emp.hora_entrada.minute
                    neto = max(0, (wd.duration_minutes or 0) - 60)
                    atraso = max(0, ls.hour * 60 + ls.minute - ref)
                    row_data = [
                        day.strftime('%d-%m-%Y'), _DAY_NAMES[day.weekday()],
                        ls.strftime('%H:%M'), le.strftime('%H:%M') if le else '—',
                        '1:00', f'{neto // 60:02d}:{neto % 60:02d}', atraso,
                    ]
                else:
                    row_data = [day.strftime('%d-%m-%Y'), _DAY_NAMES[day.weekday()], '', '', '', '', '']
                groups[emp_counter]['rows'].append({
                    'data': row_data, 'emp_id': emp.id, 'date': day,
                    'no_cerro': bool(wd and wd.auto_closed),
                })

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ── Coca-Cola palette ──────────────────────────────────
        CC_RED      = 'CC0000'
        CC_DARK_RED = '8B0000'
        CC_WHITE    = 'FFFFFF'
        CC_CREAM    = 'FFF5F5'
        CC_ROW_A    = 'FFFFFF'
        CC_ROW_B    = 'FDF0F0'
        CC_WEEKEND  = 'FFE8E8'
        CC_DIM      = 'AAAAAA'
        CC_TEXT     = '1A1A1A'
        CC_BORDER   = 'E0C0C0'
        CC_LEAVE    = 'CC2222'
        CC_NOTE     = '336699'

        def _fill(c):
            return PatternFill('solid', fgColor=c)

        def _font(bold=False, color=CC_TEXT, size=10, italic=False):
            return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)

        def _border(color=CC_BORDER):
            s = Side(style='thin', color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        def _align(h='left', v='center', wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        COL_WIDTHS   = [13, 10, 14, 14, 11, 17, 14, 32]
        COL_HEADERS  = ['Fecha', 'Día', 'Hora Ingreso', 'Hora Salida',
                        'Refrigerio', 'Horas Trabajadas', 'Atrasos (min)', 'Comentario']
        INFO_LABELS  = ['Item', 'Nombre', 'Cargo', 'Haber Básico']

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        used_names = {}
        for emp_num in order:
            g = groups[emp_num]

            name = _xls_sheet_name(g['nombre'])
            if used_names.get(name):
                used_names[name] += 1
                name = (name + str(used_names[name]))[:31]
            else:
                used_names[name] = 1
            ws = wb.create_sheet(title=name)

            # ── Column widths ──────────────────────────────────
            for ci, w in enumerate(COL_WIDTHS, start=1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            # ── Employee header rows 1-4 ───────────────────────
            info_vals = [g['emp_num'], g['nombre'], g['cargo'],
                         f"Bs. {g['haber_basico']}" if g['haber_basico'] else '—']
            for ri, (lbl, val) in enumerate(zip(INFO_LABELS, info_vals), start=1):
                ca = ws.cell(row=ri, column=1, value=lbl)
                ca.fill = _fill(CC_DARK_RED)
                ca.font = _font(bold=True, color=CC_WHITE, size=9)
                ca.alignment = _align('left')
                ca.border = Border(bottom=Side(style='thin', color=CC_RED))

                cb = ws.cell(row=ri, column=2, value=val)
                cb.fill = _fill(CC_CREAM)
                cb.font = _font(bold=(ri == 2), color=CC_TEXT, size=10)
                cb.alignment = _align('left')
                cb.border = Border(bottom=Side(style='thin', color='EEC0C0'))

                for ci in range(3, 9):
                    c = ws.cell(row=ri, column=ci)
                    c.fill = _fill(CC_CREAM)
                    c.border = Border(bottom=Side(style='thin', color='EEC0C0'))
                ws.row_dimensions[ri].height = 18

            # ── Row 5: thin separator ──────────────────────────
            for ci in range(1, 9):
                c = ws.cell(row=5, column=ci)
                c.fill = _fill(CC_RED)
            ws.row_dimensions[5].height = 4

            # ── Row 6: column headers ──────────────────────────
            for ci, lbl in enumerate(COL_HEADERS, start=1):
                c = ws.cell(row=6, column=ci, value=lbl)
                c.fill = _fill(CC_RED)
                c.font = _font(bold=True, color=CC_WHITE, size=10)
                c.alignment = _align('center')
                c.border = Border(
                    left=Side(style='thin', color='AA0000'),
                    right=Side(style='thin', color='AA0000'),
                    bottom=Side(style='medium', color=CC_DARK_RED),
                )
            ws.row_dimensions[6].height = 22

            # ── Data rows from row 7 ───────────────────────────
            for i, row_obj in enumerate(g['rows']):
                r = 7 + i
                has_wd   = bool(row_obj['data'][2])  # hora_ingreso not empty
                is_wkend = row_obj['date'].weekday() >= 5
                row_bg   = CC_WEEKEND if is_wkend else CC_ROW_A
                txt_col  = CC_DIM if not has_wd else CC_TEXT

                for ci, val in enumerate(row_obj['data'], start=1):
                    cell = ws.cell(row=r, column=ci, value=val)
                    cell.fill = _fill(row_bg)
                    cell.font = _font(color=txt_col, size=9)
                    cell.border = _border()
                    cell.alignment = _align('left' if ci <= 2 else 'center')
                ws.row_dimensions[r].height = 16

                # Refrigerio: only if workday exists
                ws.cell(row=r, column=5).value = '1:00' if has_wd else ''

                # Comment col 8
                leaves_c = xls_leaves_lookup.get((row_obj['emp_id'], row_obj['date']), [])
                notes_c  = xls_notes_lookup.get(row_obj['date'], [])
                marks    = ['No cerró jornada'] if row_obj.get('no_cerro') else []
                comment  = ' · '.join(marks + leaves_c + notes_c)
                cc = ws.cell(row=r, column=8, value=comment)
                cc.fill = _fill(row_bg)
                cc.border = _border()
                cc.alignment = _align('left', wrap=True)
                if comment:
                    cc.font = _font(
                        color=CC_RED if row_obj.get('no_cerro') else (CC_LEAVE if leaves_c else CC_NOTE),
                        size=9, italic=True,
                    )

            # ── Freeze panes below headers ─────────────────────
            ws.freeze_panes = 'A7'

            # ── Thin red bottom border on last data row ────────
            last_r = 7 + len(g['rows']) - 1
            if last_r >= 7:
                for ci in range(1, 9):
                    ws.cell(row=last_r, column=ci).border = Border(
                        left=Side(style='thin', color=CC_BORDER),
                        right=Side(style='thin', color=CC_BORDER),
                        top=Side(style='thin', color=CC_BORDER),
                        bottom=Side(style='medium', color=CC_RED),
                    )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'reporte-asistencia-{label}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ════════════════════════════════════════════════════════════════════════════
#  CERTIFICADO DE PAGO — rellena templates/plantilla.xlsx con los datos reales
# ════════════════════════════════════════════════════════════════════════════

CERT_TEMPLATE_PATH = os.path.join(settings.BASE_DIR, 'templates', 'plantilla.xlsx')
CERT_EXCHANGE_RATE = 0.461310893326238

# Geometría de la plantilla (creada para 12 empleados y un mes de 30 días)
CERT_T_DAYS       = 30   # filas de días en las hojas de empleado (10..39)
CERT_T_EMPLOYEES  = 12   # filas de empleados en PL (14..25) y hojas '1'..'12'
CERT_E_FIRST_ROW  = 10   # primera fila de datos en hoja de empleado
CERT_PL_FIRST_ROW = 14   # primera fila de empleados en PL
CERT_MAX_DAYS     = 31   # columnas de días disponibles en PL (F..AJ)
# Filas de empleados por ciudad en las planillas de alimentación de la plantilla
CERT_T_CATERING   = {'LPZ': 3, 'CBA': 3, 'SCZ': 6}
CERT_DATE_FMT     = 'dd-mm-yyyy'


def _cert_date_cell(ws, row, col, value):
    """Escribe una fecha con formato dd-mm-yyyy."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.number_format = CERT_DATE_FMT
    return cell


def _cert_copy_row_style(ws, src_r, dst_r, max_col):
    from copy import copy
    for c in range(1, max_col + 1):
        ws.cell(row=dst_r, column=c)._style = copy(ws.cell(row=src_r, column=c)._style)
    src_h = ws.row_dimensions[src_r].height
    if src_h is not None:
        ws.row_dimensions[dst_r].height = src_h


def _cert_resize_rows(ws, first_row, old_count, new_count, max_col):
    """Inserta/elimina filas para que el bloque que empieza en first_row pase de
    old_count a new_count filas. openpyxl no desplaza merges ni alturas de fila
    al insertar/eliminar, así que se reubican a mano."""
    n = new_count - old_count
    if n == 0:
        return
    boundary = first_row + old_count  # primera fila debajo del bloque original

    moved = []
    for m in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = m.bounds
        if min_r >= boundary:
            moved.append((min_c, min_r, max_c, max_r))
            ws.unmerge_cells(start_row=min_r, start_column=min_c,
                             end_row=max_r, end_column=max_c)

    if n > 0:
        ws.insert_rows(boundary, n)
        for r in range(boundary, boundary + n):
            _cert_copy_row_style(ws, first_row, r, max_col)
    else:
        ws.delete_rows(first_row + new_count, -n)
        for r in range(first_row + new_count, boundary):
            if r in ws.row_dimensions:
                ws.row_dimensions[r].height = None

    for min_c, min_r, max_c, max_r in moved:
        ws.merge_cells(start_row=min_r + n, start_column=min_c,
                       end_row=max_r + n, end_column=max_c)

    heights = sorted(((r, d.height) for r, d in ws.row_dimensions.items()
                      if r >= boundary and d.height is not None),
                     reverse=(n > 0))
    for r, h in heights:
        ws.row_dimensions[r].height = None
        ws.row_dimensions[r + n].height = h


def _cert_day_cell(ws, row, d, value):
    """Escribe la celda del día d (1..31) en una fila de la PL. La columna del
    día 31 (AJ) está vacía en la plantilla base: hereda el estilo del día 30."""
    col = 5 + d
    if d == 31:
        from copy import copy
        ws.cell(row=row, column=col)._style = copy(ws.cell(row=row, column=col - 1)._style)
    # asignación directa: ws.cell(value=None) NO limpia la celda
    ws.cell(row=row, column=col).value = value


def _cert_fill_employee_sheet(ws, item, emp, dates, wd_lookup, leaves_lookup, notes_lookup):
    """Rellena una hoja de empleado de la plantilla con los días del período."""
    first = CERT_E_FIRST_ROW
    ndays = len(dates)
    _cert_resize_rows(ws, first, CERT_T_DAYS, ndays, max_col=8)

    ws['B4'] = item
    ws['B5'] = emp.full_name
    ws['B6'] = emp.cargo or ''
    ws['B7'] = float(emp.haber_basico) if emp.haber_basico else 0

    for i, day in enumerate(dates):
        r = first + i
        wd = wd_lookup.get((emp.id, day))
        _cert_date_cell(ws, r, 1, day)
        ws.cell(row=r, column=2, value=f'=+A{r}')
        if wd:
            ls = timezone.localtime(wd.start_time)
            le = timezone.localtime(wd.end_time) if wd.end_time else None
            hora_in, hora_out, refrigerio = _time(ls.hour, ls.minute), le and _time(le.hour, le.minute), 1
        else:
            hora_in = hora_out = refrigerio = None
        # asignación directa: ws.cell(value=None) NO limpia la celda
        ws.cell(row=r, column=3).value = hora_in
        ws.cell(row=r, column=4).value = hora_out
        ws.cell(row=r, column=5).value = refrigerio
        ws.cell(row=r, column=6, value=f'=+(D{r}-C{r})*24-E{r}')
        ws.cell(row=r, column=7,
                value=f'=+IF(C{r}>0.333333333333333,+(C{r}-0.333333333333333)*24*60,"0")')
        marks = ['No cerró jornada'] if (wd and wd.auto_closed) else []
        obs = ' · '.join(marks + leaves_lookup.get((emp.id, day), []) + notes_lookup.get(day, []))
        ws.cell(row=r, column=8).value = obs or None

    last = first + ndays - 1
    ws.cell(row=last + 1, column=3, value=f'=+COUNT(C{first}:C{last})')
    ws.cell(row=last + 2, column=3, value=f'=SUM(F{first}:F{last})')
    ws.cell(row=last + 3, column=3, value=f'=SUM(G{first}:G{last})')


def _cert_fill_pl_sheet(ws, employees, dates, period_value):
    """Rellena la hoja PL (certificado consolidado) de la plantilla."""
    n_emp = len(employees)
    ndays = len(dates)
    first = CERT_PL_FIRST_ROW
    start_val = dates[0]

    _cert_date_cell(ws, 1, 39, _local_now().date())  # AM1
    ws['Y3'] = period_value

    def _fill_day_header(row):
        for d in range(1, CERT_MAX_DAYS + 1):
            _cert_day_cell(ws, row, d, dates[d - 1].day if d <= ndays else None)

    def _fill_person_row(r, item, emp, pu=None):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=emp.full_name)
        ws.cell(row=r, column=3, value=emp.cargo or '')
        _cert_date_cell(ws, r, 4, start_val)
        if pu is None:
            ws.cell(row=r, column=5, value=float(emp.haber_basico) if emp.haber_basico else 0)
        else:
            ws.cell(row=r, column=5, value=pu)
        for d in range(1, CERT_MAX_DAYS + 1):
            v = None
            if d <= ndays:
                v = f'=+IF(\'{item}\'!F{CERT_E_FIRST_ROW + d - 1}>1,"SI","NO")'
            _cert_day_cell(ws, r, d, v)
        ws.cell(row=r, column=37, value=f'=+COUNTIF(F{r}:AJ{r},"SI")')

    # ── Planilla de personal (plantilla: filas 14..25, subtotal 26) ──
    _cert_resize_rows(ws, first, CERT_T_EMPLOYEES, n_emp, max_col=39)
    _fill_day_header(first - 1)
    for idx, emp in enumerate(employees, start=1):
        r = first + idx - 1
        _fill_person_row(r, idx, emp)
        ws.cell(row=r, column=38, value=f'=+AK{r}/$AB$5*E{r}')
        ws.cell(row=r, column=39, value=f'=+AL{r}/{CERT_EXCHANGE_RATE}')
    pl_sub_r = first + n_emp
    ws.cell(row=pl_sub_r, column=39,
            value=f'=SUM(AM{first}:AM{pl_sub_r - 1})' if n_emp else 0)

    shift = n_emp - CERT_T_EMPLOYEES

    # ── Pasajes y viáticos / equipos (plantilla: filas 32..37) ───────
    t_first = 32 + shift
    base = dates[0]
    dim = _cal.monthrange(base.year, base.month)[1]
    ida = _date(base.year, base.month, min(22, dim))
    reg = _date(base.year, base.month, min(23, dim))
    for off in (1, 4):  # filas con fechas de viaje (33 y 36 en la plantilla)
        _cert_date_cell(ws, t_first + off, 4, ida)
        _cert_date_cell(ws, t_first + off, 5, reg)
    for i in range(5):
        r = t_first + i
        ws.cell(row=r, column=14, value=f'=+F{r}*H{r}' if i == 4 else f'=+K{r}*H{r}')
        ws.cell(row=r, column=16, value=f'=+N{r}*1.25')
    travel_sub_r = t_first + 5
    ws.cell(row=travel_sub_r, column=16, value=f'=SUM(P{t_first}:S{t_first + 4})')
    for off in (0, 2):  # filas de equipos (32 y 34 en la plantilla)
        r = t_first + off
        ws.cell(row=r, column=39, value=f'=+AL{r}*AJ{r}')
    eq_sub_r = t_first + 4
    ws.cell(row=eq_sub_r, column=39, value=f'=SUM(AM{t_first}:AM{t_first + 3})')

    # ── Planillas de alimentación por ciudad ─────────────────────────
    def _fill_catering(title_r, old_count, city_employees):
        days_r = title_r + 3
        first_emp = title_r + 4
        count = len(city_employees)
        _cert_resize_rows(ws, first_emp, old_count, count, max_col=39)
        _fill_day_header(days_r)
        for i, (item, emp) in enumerate(city_employees):
            r = first_emp + i
            _fill_person_row(r, item, emp, pu=25)
            ws.cell(row=r, column=38, value=f'=+AK{r}*E{r}')
            ws.cell(row=r, column=39, value=f'=+AL{r}*$AM${title_r}')
        sub_r = first_emp + count
        ws.cell(row=sub_r, column=39,
                value=f'=SUM(AM{first_emp}:AM{sub_r - 1})' if count else 0)
        return sub_r

    # Las planillas de alimentación agrupan por Employee.ciudad; los empleados
    # con CIUDAD_NONE ('Sin catering') no aparecen en ninguna
    by_city = {code: [(i, e) for i, e in enumerate(employees, start=1) if e.ciudad == code]
               for code in (Employee.CIUDAD_LPZ, Employee.CIUDAD_CBA, Employee.CIUDAD_SCZ)}

    lpz_sub_r = _fill_catering(39 + shift, CERT_T_CATERING[Employee.CIUDAD_LPZ],
                               by_city[Employee.CIUDAD_LPZ])
    cba_sub_r = _fill_catering(lpz_sub_r + 2, CERT_T_CATERING[Employee.CIUDAD_CBA],
                               by_city[Employee.CIUDAD_CBA])
    scz_sub_r = _fill_catering(cba_sub_r + 2, CERT_T_CATERING[Employee.CIUDAD_SCZ],
                               by_city[Employee.CIUDAD_SCZ])

    # ── Total facturación y "Son:" ───────────────────────────────────
    total_r = scz_sub_r + 2
    ws.cell(row=total_r, column=38, value=(
        f'=+AM{pl_sub_r}+P{travel_sub_r}+AM{eq_sub_r}'
        f'+AM{lpz_sub_r}+AM{cba_sub_r}+AM{scz_sub_r}'
    ))
    ws.cell(row=total_r + 2, column=23,
            value=f'="Son : " & TEXT(AL{total_r},"#,##0.00") & " Bolivianos"')


class CertificadoExportView(APIView):
    """Descarga el Certificado de Pago: rellena templates/plantilla.xlsx con
    los datos del mes o rango de fechas seleccionado. Solo superusers."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)
        if not executive.is_executive or not request.user.is_superuser:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        local_now = _local_now()
        filter_mode = request.query_params.get('mode', 'month')

        if filter_mode == 'range':
            try:
                from_date = _date.fromisoformat(request.query_params.get('from', ''))
                to_date   = _date.fromisoformat(request.query_params.get('to', ''))
                if to_date < from_date:
                    to_date = from_date
            except (ValueError, TypeError):
                filter_mode = 'month'

        if filter_mode == 'month':
            try:
                year  = int(request.query_params.get('year',  local_now.year))
                month = int(request.query_params.get('month', local_now.month))
                if not (1 <= month <= 12):
                    raise ValueError
            except (ValueError, TypeError):
                year, month = local_now.year, local_now.month
            from_date = _date(year, month, 1)
            to_date   = _date(year, month, _cal.monthrange(year, month)[1])
            period_value = from_date  # la celda PERIODO tiene formato 'mmmm yyyy'
            label = f'{year}-{month:02d}'
        else:
            # La PL solo tiene 31 columnas de días
            if (to_date - from_date).days >= CERT_MAX_DAYS:
                to_date = from_date + timedelta(days=CERT_MAX_DAYS - 1)
            period_value = f'{from_date.strftime("%d-%m-%Y")} — {to_date.strftime("%d-%m-%Y")}'
            label = f'{from_date.strftime("%d-%m-%Y")}__{to_date.strftime("%d-%m-%Y")}'

        dates = []
        d = from_date
        while d <= to_date:
            dates.append(d)
            d += timedelta(days=1)

        first_dt = timezone.make_aware(_datetime.combine(from_date, _time(0, 0, 0)))
        last_dt  = timezone.make_aware(_datetime.combine(to_date,   _time(23, 59, 59)))

        employees = list(Employee.objects
                         .filter(is_active=True, is_executive=False)
                         .order_by(models.F('item_number').asc(nulls_last=True), 'full_name'))

        wd_lookup = {}
        for wd in Workday.objects.filter(
            employee__in=employees,
            status__in=[Workday.STATUS_COMPLETED, Workday.STATUS_INCOMPLETE],
            start_time__gte=first_dt, start_time__lte=last_dt,
        ).select_related('employee'):
            ls = timezone.localtime(wd.start_time)
            wd_lookup[(wd.employee_id, ls.date())] = wd

        leave_labels = dict(EmployeeLeave.TYPE_CHOICES)
        leaves_lookup = {}
        for lv in EmployeeLeave.objects.filter(
            employee__in=employees,
            start_date__lte=to_date, end_date__gte=from_date,
        ):
            d = lv.start_date
            while d <= lv.end_date:
                if from_date <= d <= to_date:
                    txt = leave_labels.get(lv.leave_type, lv.leave_type)
                    if lv.note:
                        txt += f' ({lv.note})'
                    leaves_lookup.setdefault((lv.employee_id, d), []).append(txt)
                d += timedelta(days=1)

        note_labels = dict(CalendarNote.TYPE_CHOICES)
        notes_lookup = {}
        for note in CalendarNote.objects.filter(date__gte=from_date, date__lte=to_date):
            lbl = note_labels.get(note.note_type, note.note_type)
            notes_lookup.setdefault(note.date, []).append(f'{lbl}: {note.text}')

        import openpyxl
        wb = openpyxl.load_workbook(CERT_TEMPLATE_PATH)

        # La plantilla trae hojas '1'..'12': se quitan las que sobran o se
        # clonan más si hay empleados adicionales
        n_emp = len(employees)
        for s in range(n_emp + 1, CERT_T_EMPLOYEES + 1):
            wb.remove(wb[str(s)])
        for s in range(CERT_T_EMPLOYEES + 1, n_emp + 1):
            new_ws = wb.copy_worksheet(wb['1'])
            new_ws.title = str(s)

        _cert_fill_pl_sheet(wb['PL'], employees, dates, period_value)
        for idx, emp in enumerate(employees, start=1):
            _cert_fill_employee_sheet(wb[str(idx)], idx, emp, dates,
                                      wd_lookup, leaves_lookup, notes_lookup)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'certificado-{label}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ResetMobileDeviceView(APIView):
    """Ejecutivo desvincula el dispositivo móvil de un empleado."""
    permission_classes = [IsAuthenticated]

    def post(self, request, employee_id):
        try:
            executive = request.user.employee
        except Exception:
            return Response({'error': 'Perfil no encontrado'}, status=404)

        if not executive.is_executive:
            return Response({'error': 'Acceso no autorizado'}, status=403)

        try:
            emp = Employee.objects.get(id=employee_id, is_active=True, is_executive=False)
        except Employee.DoesNotExist:
            return Response({'error': 'Empleado no encontrado'}, status=404)

        emp.mobile_device_id = ''
        emp.save(update_fields=['mobile_device_id'])
        return Response({'status': 'ok'})
